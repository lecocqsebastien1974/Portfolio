#!/usr/bin/env python
import os
import sys
import django

# Setup Django
sys.path.append('/app')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'portfolio_backend.settings')
django.setup()

from portfolios.models import Signaletique, AssetCategory
import openpyxl

# Charger le fichier Excel
wb = openpyxl.load_workbook('/app/signaletique_valides.xlsx')
ws = wb.active

# Récupérer les en-têtes
headers = [cell.value for cell in ws[1]]

# Parcourir les lignes
rows = list(ws.iter_rows(min_row=2, values_only=True))
count = 0
updated = 0
created = 0

for row in rows:
    row_data = {header: value for header, value in zip(headers, row)}
    
    isin = row_data.get('Isin')
    titre = row_data.get('Nom', '')
    type_instr = row_data.get("Type d'instr")
    classe = row_data.get("Classe d'actifs")
    
    if not isin:
        continue
    
    # Gérer la catégorie
    categorie_instance = None
    if classe:
        categorie_instance, _ = AssetCategory.objects.get_or_create(
            name=str(classe).strip().capitalize()
        )
    
    # Mettre à jour ou créer
    sig, is_created = Signaletique.objects.update_or_create(
        isin=isin,
        defaults={
            'code': f'SIG_{isin}',
            'titre': str(titre)[:500],
            'categorie': categorie_instance,
            'categorie_text': str(classe) if classe else None,
            'statut': str(type_instr) if type_instr else None,
            'donnees_supplementaires': row_data
        }
    )
    
    count += 1
    if is_created:
        created += 1
    else:
        updated += 1

print(f"✅ Import terminé !")
print(f"📊 Total traité : {count}")
print(f"➕ Créées : {created}")
print(f"🔄 Mises à jour : {updated}")
