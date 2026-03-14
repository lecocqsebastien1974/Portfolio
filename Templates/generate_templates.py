"""
Génère tous les templates d'import pour le projet Portfolio.
Exécuter avec : python generate_templates.py
(nécessite openpyxl : pip install openpyxl)
"""

import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def style_header(ws, nb_cols):
    """Applique un style titre sur la première ligne."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color='FFFFFF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx in range(1, nb_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30


def add_example_row(ws, values):
    """Ajoute une ligne d'exemple en italique gris."""
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    example_font = Font(italic=True, color="808080", size=10)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.fill = example_fill
        cell.font = example_font


def auto_width(ws):
    """Ajuste automatiquement la largeur des colonnes."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)


# ─────────────────────────────────────────────────────────────────────────────
# 1. SIGNALÉTIQUE
# ─────────────────────────────────────────────────────────────────────────────
def create_signaletique():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signaletique"

    headers = [
        "Type d'instr", "Isin", "Classe d'actifs", "Nom", "Symbole",
        "Banques Dispo", "Devise", "Taux", "Date de fin", "Qualité credit",
        "TER", "Cap/Dis", "ESG", "Replication", "Taille du Fonds",
        "Positions", "Couverture de change",
        # Géographie
        "USA", "Japon", "Grande Bretagne", "Canada",
        "Pays Emergeants Hors Chine et Japon", "Australie", "Suède",
        "Suisse", "Chine", "Israel", "Allemagne", "Nouvelle Zelande",
        "Pays-Bas", "Irlande", "Espagne", "Italie", "France", "Autre Pays",
        # Secteurs
        "Etats", "Industrie", "Finance", "Consommation Cyclique", "Technologie",
        "Santé", "Consommation Defensive", "Communication", "Immobilier",
        "Matières Premières", "Energie", "Service Publiques",
        "Services de consommation", "Autre Secteur",
        # Emetteurs
        "Etats2", "Banque Emetteur", "Entreprises", "Autre Emetteur", "CodeBank",
        # Toujours en dernière position
        "Frequence Coupon",
    ]

    ws.append(headers)
    style_header(ws, len(headers))

    example = [
        "ETF", "FR0010315770", "Actions", "Lyxor CAC 40 DR UCITS ETF", "CAC",
        "Bourse", "EUR", "", "", "BBB",
        "0,20%", "Capitalisation", "Oui", "Physique", "1 000 M€",
        "40", "Non",
        # Géo
        "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0", "0",
        "0", "0", "0", "100", "0",
        # Secteurs
        "0", "20", "15", "10", "15", "10", "5", "5", "5", "5", "4", "1", "0", "5",
        # Emetteurs
        "0", "0", "100", "0", "CAC",
        # Frequence Coupon
        "",
    ]
    add_example_row(ws, example)
    auto_width(ws)
    out = os.path.join(OUTPUT_DIR, "signaletique_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. TRANSACTIONS
# ─────────────────────────────────────────────────────────────────────────────
def create_transactions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Type", "Isin", "Quantité", "Prix unitaire", "Devise", "Portefeuille"]
    ws.append(headers)
    style_header(ws, len(headers))

    add_example_row(ws, ["2026-01-15", "ACHAT", "FR0010315770", 10, 34.50, "EUR", "MonPortefeuille"])
    auto_width(ws)

    # Note sur les valeurs acceptées
    ws.cell(row=3, column=1, value="← Format AAAA-MM-JJ")
    ws.cell(row=3, column=2, value="← ACHAT ou VENTE")
    ws.cell(row=3, column=3, value="← 12 caractères exactement")
    for col in range(1, 4):
        ws.cell(row=3, column=col).font = Font(italic=True, color="C0392B", size=9)

    out = os.path.join(OUTPUT_DIR, "transactions_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 3. CASH
# ─────────────────────────────────────────────────────────────────────────────
def create_cash():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash"

    headers = ["Portefeuille", "Banque", "Montant", "Devise", "Date", "Commentaire"]
    ws.append(headers)
    style_header(ws, len(headers))

    add_example_row(ws, ["MonPortefeuille", "BNP Paribas", 5000.00, "EUR", "2026-01-01", "Solde initial"])
    auto_width(ws)

    out = os.path.join(OUTPUT_DIR, "cash_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 4. PORTEFEUILLES
# ─────────────────────────────────────────────────────────────────────────────
def create_portefeuilles():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portefeuilles"

    headers = ["Nom", "Description", "Type", "Courtier", "Devise", "Date ouverture", "Couleur"]
    ws.append(headers)
    style_header(ws, len(headers))

    add_example_row(ws, ["MonPortefeuille", "Portefeuille principal", "PEA", "Saxo Bank", "EUR", "2020-01-01", "#2980B9"])
    auto_width(ws)

    out = os.path.join(OUTPUT_DIR, "portefeuilles_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 5. PORTEFEUILLE CIBLE
# ─────────────────────────────────────────────────────────────────────────────
def create_portefeuille_cible():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Portefeuille cible"

    headers = ["Portefeuille", "ISIN", "Titre", "Ratio"]
    ws.append(headers)
    style_header(ws, len(headers))

    add_example_row(ws, ["Cible 60/40", "FR0010315770", "Lyxor CAC 40", 0.60])
    ws.append(["Cible 60/40", "FR0010251744", "Lyxor Euro Gov Bond", 0.40])
    auto_width(ws)

    ws.cell(row=4, column=4, value="← Ratio entre 0 et 1 (ex: 0.25 = 25%)")
    ws.cell(row=4, column=4).font = Font(italic=True, color="C0392B", size=9)

    out = os.path.join(OUTPUT_DIR, "portefeuille_cible_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 6. UTILISATEURS
# ─────────────────────────────────────────────────────────────────────────────
def create_utilisateurs():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    headers = ["username", "email", "first_name", "last_name",
               "is_staff", "is_superuser", "is_active", "password_hash"]
    ws.append(headers)
    style_header(ws, len(headers))

    add_example_row(ws, [
        "john.doe", "john.doe@example.com", "John", "Doe",
        False, False, True,
        "pbkdf2_sha256$600000$..."
    ])

    ws.cell(row=3, column=8, value="← Hash Django (pbkdf2_sha256$...) copié depuis une sauvegarde")
    ws.cell(row=3, column=8).font = Font(italic=True, color="C0392B", size=9)
    auto_width(ws)

    out = os.path.join(OUTPUT_DIR, "utilisateurs_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 7. PRIX HISTORIQUE (par titre)
# ─────────────────────────────────────────────────────────────────────────────
def create_prix_historique():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prix historique"

    headers = ["Date", "Cours", "Devise", "Source", "Symbole"]
    ws.append(headers)
    style_header(ws, len(headers))

    add_example_row(ws, ["2026-01-15", 34.50, "EUR", "Manuel", "CAC"])

    ws.cell(row=3, column=1, value="← Format AAAA-MM-JJ")
    ws.cell(row=3, column=4, value="← Manuel, Koala, Bonobo, etc.")
    for col in [1, 4]:
        ws.cell(row=3, column=col).font = Font(italic=True, color="C0392B", size=9)

    auto_width(ws)

    out = os.path.join(OUTPUT_DIR, "prix_historique_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 8. KOALA (source de cotation)
# ─────────────────────────────────────────────────────────────────────────────
def create_koala():
    """
    Le module Koala lit spécifiquement :
      - Colonne D (index 4) = Symbole
      - Colonne H (index 8) = Cours
    Les autres colonnes sont ignorées.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Positions"  # La feuille doit s'appeler Positions

    # 8 colonnes obligatoires ; D et H sont les seules lues par le module
    headers = ["Compte", "Catégorie", "Secteur", "Symbole", "Nom", "Quantité", "Devise", "Cours"]
    ws.append(headers)
    style_header(ws, len(headers))

    # Mise en évidence des colonnes critiques D et H
    highlight = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold_orange = Font(bold=True, color="C0392B", size=10)
    for col_letter in ["D", "H"]:
        cell = ws[f"{col_letter}1"]
        cell.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)

    add_example_row(ws, ["COMPTE1", "Actions", "Technologie", "AAPL", "Apple Inc.", 10, "USD", 182.50])
    for col_letter in ["D", "H"]:
        ws[f"{col_letter}2"].fill = highlight

    # Note explicative
    ws.cell(row=3, column=4, value="⬆ Colonne D : SYMBOLE (lu par le module)")
    ws.cell(row=3, column=8, value="⬆ Colonne H : COURS (lu par le module)")
    for col in [4, 8]:
        ws.cell(row=3, column=col).font = Font(italic=True, color="C0392B", size=9)

    auto_width(ws)

    out = os.path.join(OUTPUT_DIR, "koala_template.xlsx")
    wb.save(out)
    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# 9. BONOBO (source de cotation, CSV)
# ─────────────────────────────────────────────────────────────────────────────
def create_bonobo():
    """
    Le module Bonobo lit un CSV délimité par ';'.
    Colonnes retenues (positions fixes) :
      - Col 3 (index 2) : ISIN
      - Col 5 (index 4) : Cours
      - Col 6 (index 5) : Devise
    """
    out = os.path.join(OUTPUT_DIR, "bonobo_template.csv")
    headers = ["Nom", "Type", "ISIN", "Description", "Cours", "Devise"]
    example = ["Lyxor CAC 40", "ETF", "FR0010315770", "CAC 40 tracker", "34,50", "EUR"]

    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(headers)
        writer.writerow(example)
        # Note dans le fichier
        writer.writerow(["", "", "⬆ Col 3 : ISIN (lu)", "", "⬆ Col 5 : Cours (lu)", "⬆ Col 6 : Devise (lu)"])

    print(f"  ✓ {os.path.basename(out)}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\nGénération des templates dans : {OUTPUT_DIR}\n")
    create_signaletique()
    create_transactions()
    create_cash()
    create_portefeuilles()
    create_portefeuille_cible()
    create_utilisateurs()
    create_prix_historique()
    create_koala()
    create_bonobo()
    print("\nTous les templates ont été générés avec succès !")
