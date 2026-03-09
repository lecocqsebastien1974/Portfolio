#!/usr/bin/env python
import openpyxl

wb = openpyxl.load_workbook('/app/achatall.xlsx')
ws = wb.active

print(f"Lignes: {ws.max_row}")
print(f"Colonnes: {ws.max_column}")
print("\nEn-têtes:")
print([cell.value for cell in ws[1]])

print("\nPremières transactions:")
for row in list(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
    print(f"  Date: {row[0]}, Sens: {row[1]}, ISIN: {row[2]}")
    print(f"    Qty: {row[3]}, Prix: {row[4]}, Devise: {row[5]}, Portfolio: {row[6]}")
