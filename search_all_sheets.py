#!/usr/bin/env python
import openpyxl

# Charger le fichier Excel
wb = openpyxl.load_workbook('/app/achat_ETF_et_fonds_oblig2.xlsx')

print(f"Feuilles disponibles: {wb.sheetnames}")
print()

target_isins = ['FR0010149179', 'GB00B00FHZ82', 'LU0171289902']

# Parcourir toutes les feuilles
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"Feuille: {sheet_name}")
    print('='*80)
    
    # Obtenir les en-têtes
    headers = [cell.value for cell in ws[1]]
    print(f"En-têtes: {headers[:10]}")  # Premiers 10
    
    # Trouver la colonne ISIN
    isin_col_idx = None
    for idx, header in enumerate(headers):
        if header and 'isin' in str(header).lower():
            isin_col_idx = idx
            break
    
    if isin_col_idx is None:
        print("⚠️  Pas de colonne ISIN trouvée")
        continue
    
    print(f"Colonne ISIN: index {isin_col_idx} ('{headers[isin_col_idx]}')")
    
    # Chercher les ISINs
    found_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        isin = row[isin_col_idx] if isin_col_idx < len(row) else None
        if isin:
            isin_str = str(isin).strip()
            if isin_str in target_isins:
                print(f"\n✓ ISIN trouvé à la ligne {row_idx}: {isin}")
                for idx, (header, value) in enumerate(zip(headers, row)):
                    if value:
                        print(f"  {header}: {value}")
                found_count += 1
    
    if found_count == 0:
        print("  → Aucun ISIN cible trouvé dans cette feuille")
        # Afficher tous les ISINs de cette feuille
        all_isins = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            isin = row[isin_col_idx] if isin_col_idx < len(row) else None
            if isin and str(isin).strip():
                all_isins.add(str(isin).strip())
        
        if all_isins:
            print(f"\n  ISINs présents ({len(all_isins)}):")
            for isin in sorted(list(all_isins))[:20]:  # Max 20 premiers
                print(f"    - {isin}")
