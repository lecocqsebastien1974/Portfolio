#!/usr/bin/env python
import os
import sys
import django

# Setup Django
sys.path.append('/app')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'portfolio_backend.settings')
django.setup()

from portfolios.models import Signaletique, AssetCategory
from datetime import datetime, date
import openpyxl

# ISINs à traiter
target_isins = ['FR0010149179', 'GB00B00FHZ82']

# Charger le fichier
wb = openpyxl.load_workbook('/app/Signalétique Template gb et carmi.xlsx')
ws = wb['Positions']

# Obtenir les en-têtes
headers = [cell.value for cell in ws[1]]

# Trouver l'index de la colonne ISIN
isin_col_idx = None
for idx, h in enumerate(headers):
    if h and 'isin' in str(h).lower():
        isin_col_idx = idx
        break

if isin_col_idx is None:
    print("❌ Colonne ISIN introuvable")
    sys.exit(1)

print(f"✓ Colonne ISIN trouvée à l'index {isin_col_idx}")
print(f"\n{'='*80}")
print("Import des signalétiques GB et Carmignac:")
print('='*80)

count = 0
updated = 0
created = 0

# Parcourir les lignes
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    isin = row[isin_col_idx] if isin_col_idx < len(row) else None
    
    if not isin or str(isin).strip() not in target_isins:
        continue
    
    # Créer un dictionnaire des données
    row_data = {}
    for idx, (header, value) in enumerate(zip(headers, row)):
        if value is not None and str(value).strip():
            # Ignorer les formules Excel
            if isinstance(value, str) and value.startswith('='):
                continue
            # Convertir les dates en string pour JSON
            if isinstance(value, (datetime, date)):
                row_data[header] = value.isoformat()
            else:
                row_data[header] = value
    
    isin_str = str(isin).strip()
    titre = row_data.get('Nom', '') or row_data.get('Nom Long', '')
    type_instr = row_data.get("Type d'instr", '')
    classe = row_data.get("Classe d'actifs", '')
    
    print(f"\n{count+1}. ISIN: {isin_str}")
    print(f"   Titre: {titre}")
    print(f"   Type: {type_instr}")
    print(f"   Classe: {classe}")
    
    # Gérer la catégorie
    categorie_instance = None
    if classe:
        categorie_instance, _ = AssetCategory.objects.get_or_create(
            name=str(classe).strip().capitalize()
        )
    
    # Mettre à jour ou créer
    try:
        sig, is_created = Signaletique.objects.update_or_create(
            isin=isin_str,
            defaults={
                'code': f'SIG_{isin_str}',
                'titre': str(titre)[:500] if titre else f'[À compléter] {isin_str}',
                'categorie': categorie_instance,
                'categorie_text': str(classe) if classe else None,
                'statut': str(type_instr) if type_instr else None,
                'donnees_supplementaires': row_data
            }
        )
        
        count += 1
        if is_created:
            created += 1
            print(f"   ✅ Créé")
        else:
            updated += 1
            print(f"   🔄 Mis à jour")
    
    except Exception as e:
        print(f"   ❌ Erreur: {e}")

print(f"\n{'='*80}")
print(f"✅ Import terminé !")
print(f"📊 Total traité : {count}")
print(f"➕ Créées : {created}")
print(f"🔄 Mises à jour : {updated}")
print('='*80)
