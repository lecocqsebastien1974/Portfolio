from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework import status
from django.db import models
import openpyxl
import pandas as pd
import csv
import os
from datetime import date, datetime
from decimal import Decimal
from django.conf import settings
from .models import Signaletique, ImportLog, TargetPortfolio, RealPortfolio, Transaction, AssetCategory, Cash
from .serializers import (
    SignaletiqueSerializer,
    ImportLogSerializer,
    TargetPortfolioSerializer,
    RealPortfolioSerializer,
    TransactionSerializer,
    CashSerializer
)
from . import file_storage

@api_view(['GET'])
def health_check(request):
    return Response({'status': 'ok', 'message': 'Portfolio API is running'})


@api_view(['GET'])
def list_categories(request):
    """Liste toutes les catégories d'actifs disponibles"""
    categories = AssetCategory.objects.all().order_by('ordre', 'name')
    return Response([
        {
            'id': cat.id,
            'name': cat.name,
            'description': cat.description,
            'color': cat.color,
            'ordre': cat.ordre
        }
        for cat in categories
    ])


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_signaletique(request):
    """Import des données de signalétique depuis un fichier Excel"""
    
    if 'file' not in request.FILES:
        return Response(
            {'error': 'Aucun fichier fourni'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    file = request.FILES['file']
    
    # Vérifier l'extension
    if not file.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'Format de fichier non supporté. Utilisez .xlsx ou .xls'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Créer un log d'import
    import_log = ImportLog.objects.create(
        type_import='signaletique',
        nom_fichier=file.name,
        statut='en_cours'
    )
    
    try:
        # Lire le fichier Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Récupérer les en-têtes (première ligne)
        headers = [cell.value for cell in ws[1]]
        
        nombre_succes = 0
        nombre_erreurs = 0
        nombre_ignores = 0
        erreurs = []

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        def sanitize_value(value):
            if isinstance(value, (datetime, date)):
                return value.isoformat()
            return value

        # Ne rien supprimer ni désactiver - l'import crée ou met à jour uniquement
        # Les signalétiques non présentes dans le fichier restent inchangées

        # Parcourir les lignes (à partir de la ligne 2)
        for row_idx, row in enumerate(rows, start=2):
            try:
                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue
                # Créer un dictionnaire avec les données de la ligne
                row_data = {
                    header: sanitize_value(value)
                    for header, value in zip(headers, row)
                }
                
                # Extraire les champs principaux (mapping adapté à Signalétique.xlsx)
                # Colonnes principales reconnaissables
                isin_raw = (
                    row_data.get('Isin') or 
                    row_data.get('ISIN') or 
                    row_data.get('isin')
                )
                isin_value = str(isin_raw).strip().upper() if isin_raw is not None else None
                if isin_value == "":
                    isin_value = None
                
                # Validation de l'ISIN (doit avoir 12 caractères)
                if isin_value and len(isin_value) != 12:
                    erreurs.append({
                        'ligne': row_idx,
                        'erreur': f'ISIN invalide (longueur {len(isin_value)} au lieu de 12): {isin_value}'
                    })
                    nombre_erreurs += 1
                    continue
                
                # Code généré à partir de ISIN ou auto-incrémenté
                if isin_value:
                    code = f"SIG_{isin_value}"
                else:
                    code = f"AUTO_{row_idx}"
                
                titre = row_data.get('Nom') or ""
                
                description = (
                    row_data.get('Description') or 
                    row_data.get('description') or 
                    ""
                )
                
                categorie = row_data.get('Classe d\'actifs') or ""
                
                statut_data = row_data.get('Type d\'instr') or ""
                
                # Gérer la catégorie d'actifs (on conserve juste le texte)
                categorie_text = str(categorie).strip().capitalize() if categorie and str(categorie).strip() else None

                # Persister dans le JSON via file_storage
                sig_dict = file_storage.upsert_signaletique(
                    code=str(code),
                    isin=isin_value,
                    titre=str(titre)[:500],
                    description=str(description) if description else None,
                    categorie_text=categorie_text,
                    statut=str(statut_data)[:100] if statut_data else None,
                    donnees_supplementaires=row_data
                )

                # Double-write vers PostgreSQL pour compatibilité FK résiduelle
                categorie_instance = None
                if categorie_text:
                    categorie_instance, _ = AssetCategory.objects.get_or_create(
                        name=categorie_text,
                        defaults={'description': 'Catégorie créée automatiquement lors de l\'import'}
                    )

                db_defaults = {
                    'code': str(code),
                    'isin': isin_value,
                    'titre': str(titre)[:500],
                    'description': str(description) if description else None,
                    'categorie': categorie_instance,
                    'categorie_text': categorie_text,
                    'statut': str(statut_data)[:100] if statut_data else None,
                    'donnees_supplementaires': row_data
                }

                if not isin_value and not str(titre).strip():
                    continue

                if isin_value:
                    Signaletique.objects.update_or_create(isin=isin_value, defaults=db_defaults)
                else:
                    Signaletique.objects.update_or_create(code=str(code), defaults=db_defaults)
                
                nombre_succes += 1
                
            except Exception as e:
                nombre_erreurs += 1
                erreurs.append({
                    'ligne': row_idx,
                    'erreur': str(e)
                })
        
        # Mettre à jour le log
        import_log.nombre_lignes = ws.max_row - 1
        import_log.nombre_succes = nombre_succes
        import_log.nombre_erreurs = nombre_erreurs
        import_log.erreurs = erreurs if erreurs else None
        import_log.statut = 'termine' if nombre_erreurs == 0 else 'termine_avec_erreurs'
        import_log.save()
        
        return Response({
            'success': True,
            'message': f'Import terminé avec succès',
            'details': {
                'fichier': file.name,
                'colonnes_detectees': headers,
                'lignes_totales': ws.max_row - 1,
                'succes': nombre_succes,
                'erreurs': nombre_erreurs,
                'ignores': nombre_ignores,
                'liste_erreurs': erreurs[:10] if erreurs else []  # Limiter à 10 erreurs
            }
        })
        
    except Exception as e:
        import_log.statut = 'erreur'
        import_log.erreurs = [{'erreur_generale': str(e)}]
        import_log.save()
        
        return Response(
            {'error': f'Erreur lors de l\'import: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET', 'POST'])
def list_signaletique(request):
    """Lister ou créer des signalétiques"""
    if request.method == 'GET':
        return Response(file_storage.get_all_signaletiques())

    elif request.method == 'POST':
        isin_raw = request.data.get('isin', '')
        isin = isin_raw.strip().upper() if isin_raw else None
        titre = request.data.get('titre', '')
        code = request.data.get('code', '')
        if isin and len(isin) != 12:
            return Response({'error': "L'ISIN doit avoir exactement 12 caractères"}, status=status.HTTP_400_BAD_REQUEST)
        if not code and isin:
            code = f'SIG_{isin}'
        if not titre:
            return Response({'error': 'Le titre est requis'}, status=status.HTTP_400_BAD_REQUEST)
        sig_dict = file_storage.upsert_signaletique(
            code=code, isin=isin, titre=titre,
            description=request.data.get('description'),
            categorie_text=request.data.get('categorie_text'),
            statut=request.data.get('statut'),
            donnees_supplementaires=request.data.get('donnees_supplementaires')
        )
        # Double-write DB
        cat = None
        cat_name = request.data.get('categorie_text')
        if cat_name:
            cat, _ = AssetCategory.objects.get_or_create(name=cat_name)
        db_defaults = {'code': code, 'isin': isin, 'titre': titre,
                       'description': request.data.get('description'),
                       'categorie': cat, 'categorie_text': cat_name,
                       'statut': request.data.get('statut'),
                       'donnees_supplementaires': request.data.get('donnees_supplementaires')}
        if isin:
            Signaletique.objects.update_or_create(isin=isin, defaults=db_defaults)
        else:
            Signaletique.objects.update_or_create(code=code, defaults=db_defaults)
        return Response(sig_dict, status=status.HTTP_201_CREATED)


@api_view(['POST'])
def clear_signaletique(request):
    # Suppression désactivée - les signalétiques ne peuvent pas être supprimées
    return Response(
        {'error': 'La suppression des signalétiques est désactivée pour préserver l\'intégrité des données'},
        status=status.HTTP_403_FORBIDDEN
    )


@api_view(['GET'])
def export_signaletique_csv(request):
    """Exporter tous les titres de la signalétique en CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="signaletique_export.csv"'
    response.write('\ufeff')  # BOM UTF-8 pour Excel

    writer = csv.writer(response)

    signaletiques = file_storage.get_all_signaletiques()
    if not signaletiques:
        writer.writerow(['Aucune donnée à exporter'])
        return response

    all_fields = set()
    for sig in signaletiques:
        ds = sig.get('donnees_supplementaires') or {}
        all_fields.update(ds.keys())

    headers = ['ID', 'Code', 'ISIN', 'Titre'] + sorted(all_fields)
    writer.writerow(headers)

    for sig in signaletiques:
        row = [sig.get('id', ''), sig.get('code', ''), sig.get('isin') or '', sig.get('titre', '')]
        ds = sig.get('donnees_supplementaires') or {}
        for field in sorted(all_fields):
            value = ds.get(field, '')
            row.append(value if value is not None else '')
        writer.writerow(row)

    return response


# ---------------------------------------------------------------------------
# Import source Koala – colonnes D (Symbole) et H (Cours) uniquement
# ---------------------------------------------------------------------------

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_koala(request):
    """
    Import source Koala : lit uniquement la colonne D (Symbole) et H (Cours)
    de la feuille active (Positions) d'un fichier .xlsx.
    Stocke l'historique dans /app/data/prix_historique_koala.json.
    Rien n'est écrit en base de données ni dans un autre fichier JSON.
    """
    if 'file' not in request.FILES:
        return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

    uploaded = request.FILES['file']
    if not uploaded.name.lower().endswith(('.xlsx', '.xls')):
        return Response({'error': 'Format non supporté. Utilisez .xlsx ou .xls'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        # On lit la première feuille (Positions)
        ws = wb.active

        date_import = date.today().isoformat()
        rows_to_add = []
        ignores = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # col D = index 3, col H = index 7
            symbole = row[3] if len(row) > 3 else None
            cours = row[7] if len(row) > 7 else None

            if symbole is None or cours is None:
                ignores += 1
                continue

            symbole_str = str(symbole).strip()
            if not symbole_str:
                ignores += 1
                continue

            try:
                cours_float = float(cours)
            except (TypeError, ValueError):
                ignores += 1
                continue

            rows_to_add.append({
                'symbole': symbole_str,
                'cours': cours_float,
                'date_import': date_import,
                'devise': 'EUR',
            })

        stats = file_storage.append_prix_koala(rows_to_add)

        return Response({
            'success': True,
            'message': f"Import source Koala terminé : {stats['ajoutes']} cours importés",
            'details': {
                'fichier': uploaded.name,
                'date_import': date_import,
                'ajoutes': stats['ajoutes'],
                'ignores': stats['ignores'] + ignores,
                'identifiants': sorted({r['symbole'] for r in rows_to_add}),
            }
        })

    except Exception as e:
        return Response({'error': f"Erreur lors de l'import Koala : {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_koala_history(request):
    """Retourne l'historique complet des cours Koala (GET /api/import/koala/history/)"""
    historique = file_storage.get_prix_historique_koala()
    return Response(historique)


# ---------------------------------------------------------------------------
# Import source Bonobo – CSV ; – col 3 (ISIN), col 5 (Cours), col 6 (Devise)
# ---------------------------------------------------------------------------

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bonobo(request):
    """
    Import source Bonobo : lit un CSV délimité par ';'.
    Colonnes retenues uniquement :
      - col 3 (index 2) : Code ISIN  → identifiant
      - col 5 (index 4) : Cours      → prix (virgule décimale acceptée)
      - col 6 (index 5) : Dev.       → devise (peut être '%')
    Tout le reste est ignoré. Rien n'est écrit en base de données.
    """
    if 'file' not in request.FILES:
        return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

    uploaded = request.FILES['file']
    if not uploaded.name.lower().endswith('.csv'):
        return Response({'error': 'Format non supporté. Utilisez un fichier .csv'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        import io
        content = uploaded.read().decode('utf-8-sig', errors='replace')
        reader = csv.reader(io.StringIO(content), delimiter=';')

        date_import = date.today().isoformat()
        rows_to_add = []
        ignores = 0
        next(reader, None)  # skip header

        for line in reader:
            if len(line) < 6:
                ignores += 1
                continue

            isin = line[2].strip()
            cours_raw = line[4].strip().replace(',', '.')
            devise = line[5].strip()

            if not isin or not cours_raw:
                ignores += 1
                continue

            try:
                cours_float = float(cours_raw)
            except ValueError:
                ignores += 1
                continue

            rows_to_add.append({
                'isin': isin,
                'cours': cours_float,
                'devise': devise if devise else 'EUR',
                'date_import': date_import,
            })

        stats = file_storage.append_prix_bonobo(rows_to_add)

        return Response({
            'success': True,
            'message': f"Import source Bonobo terminé : {stats['ajoutes']} cours importés",
            'details': {
                'fichier': uploaded.name,
                'date_import': date_import,
                'ajoutes': stats['ajoutes'],
                'ignores': stats['ignores'] + ignores,
                'identifiants': sorted({r['isin'] for r in rows_to_add}),
            }
        })

    except Exception as e:
        return Response({'error': f"Erreur lors de l'import Bonobo : {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def import_prix_startfiles(request):
    """
    Import tout-en-un : lit le .xlsx (Koala) et le .csv (Bonobo) depuis /app/startfiles,
    les importe, puis consolide vers prix_historique_titres.json.
    POST /api/import/prix-startfiles/
    """
    import glob, io

    START_DIR = '/app/startfiles'
    date_import = date.today().isoformat()
    rapport = {}

    # ── Koala (premier .xlsx trouvé) ─────────────────────────────────────────
    xlsx_files = sorted(glob.glob(os.path.join(START_DIR, '*.xlsx')))
    if xlsx_files:
        xlsx_path = xlsx_files[-1]  # le plus récent (tri alphabétique)
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            rows_to_add = []
            ignores_k = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                symbole = row[3] if len(row) > 3 else None
                cours   = row[7] if len(row) > 7 else None
                if symbole is None or cours is None:
                    ignores_k += 1
                    continue
                s = str(symbole).strip()
                if not s:
                    ignores_k += 1
                    continue
                try:
                    c = float(cours)
                except (TypeError, ValueError):
                    ignores_k += 1
                    continue
                rows_to_add.append({'symbole': s, 'cours': c, 'date_import': date_import, 'devise': 'EUR'})
            stats_k = file_storage.append_prix_koala(rows_to_add)
            rapport['koala'] = {
                'fichier': os.path.basename(xlsx_path),
                'ajoutes': stats_k['ajoutes'],
                'ignores': stats_k['ignores'] + ignores_k,
            }
        except Exception as e:
            rapport['koala'] = {'error': str(e)}
    else:
        rapport['koala'] = {'info': 'Aucun fichier .xlsx trouvé dans Start Files'}

    # ── Bonobo (premier .csv trouvé) ─────────────────────────────────────────
    csv_files = sorted(glob.glob(os.path.join(START_DIR, '*.csv')))
    if csv_files:
        csv_path = csv_files[-1]
        try:
            with open(csv_path, encoding='utf-8-sig', errors='replace') as f:
                content = f.read()
            reader = csv.reader(io.StringIO(content), delimiter=';')
            rows_to_add = []
            ignores_b = 0
            next(reader, None)  # skip header
            for line in reader:
                if len(line) < 6:
                    ignores_b += 1
                    continue
                isin      = line[2].strip()
                cours_raw = line[4].strip().replace(',', '.')
                devise    = line[5].strip()
                if not isin or not cours_raw:
                    ignores_b += 1
                    continue
                try:
                    c = float(cours_raw)
                except ValueError:
                    ignores_b += 1
                    continue
                rows_to_add.append({'isin': isin, 'cours': c, 'devise': devise or 'EUR', 'date_import': date_import})
            stats_b = file_storage.append_prix_bonobo(rows_to_add)
            rapport['bonobo'] = {
                'fichier': os.path.basename(csv_path),
                'ajoutes': stats_b['ajoutes'],
                'ignores': stats_b['ignores'] + ignores_b,
            }
        except Exception as e:
            rapport['bonobo'] = {'error': str(e)}
    else:
        rapport['bonobo'] = {'info': 'Aucun fichier .csv trouvé dans Start Files'}

    # ── Consolidation ─────────────────────────────────────────────────────────
    try:
        stats_c = file_storage.rebuild_prix_historique_titres()
        rapport['consolidation'] = stats_c
    except Exception as e:
        rapport['consolidation'] = {'error': str(e)}

    errors = [k for k, v in rapport.items() if isinstance(v, dict) and 'error' in v]
    c = rapport.get('consolidation', {})
    return Response({
        'success': len(errors) == 0,
        'message': (
            f"Koala : {rapport.get('koala', {}).get('ajoutes', 0)} cours — "
            f"Bonobo : {rapport.get('bonobo', {}).get('ajoutes', 0)} cours — "
            f"Consolidé : {c.get('titres_consolides', 0)} titres ({c.get('total_entrees', 0)} entrées)"
        ),
        'details': rapport,
    })


@api_view(['GET'])
def get_bonobo_history(request):
    """Retourne l'historique complet des cours Bonobo (GET /api/import/bonobo/history/)"""
    historique = file_storage.get_prix_historique_bonobo()
    return Response(historique)


@api_view(['POST'])
def consolidate_prices(request):
    """Consolide Koala + Bonobo → prix_historique_titres.json (POST /api/prix-historique/consolidate/)"""
    try:
        stats = file_storage.rebuild_prix_historique_titres()
        return Response({
            'success': True,
            'message': (
                f"{stats['titres_consolides']} titres consolidés "
                f"({stats['total_entrees']} entrées — "
                f"{stats['koala_matches']} Koala, {stats['bonobo_matches']} Bonobo)"
            ),
            'details': stats,
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def titre_price_history(request, isin):
    """Retourne l'historique de prix pour un titre (GET /api/prix-historique/<isin>/)"""
    data = file_storage.get_prix_titre_by_isin(isin)
    if data is None:
        return Response(
            {'error': f"Aucun historique trouvé pour {isin}"},
            status=status.HTTP_404_NOT_FOUND,
        )
    return Response(data)


@api_view(['POST'])
def backup_all_data(request):
    """Export toutes les données (cash, portefeuilles cibles, transactions, signalétique)
    vers des fichiers Excel datés dans /app/sauvegarde/.
    Les fichiers sont au même format que les modules d'import afin d'être réutilisables.
    """
    import os
    from datetime import date as date_cls
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    SAUVEGARDE_DIR = '/app/sauvegarde'
    os.makedirs(SAUVEGARDE_DIR, exist_ok=True)

    today = date_cls.today().strftime('%Y-%m-%d')
    fichiers_crees = []

    def safe_save(wb, path):
        """Supprime le fichier existant avant de sauvegarder pour éviter les erreurs de permission."""
        if os.path.exists(path):
            try:
                os.chmod(path, 0o666)
                os.remove(path)
            except OSError:
                pass  # Si toujours impossible, openpyxl écrasera ou lèvera une erreur plus claire
        try:
            wb.save(path)
        except PermissionError:
            fname = os.path.basename(path)
            raise PermissionError(
                f"Impossible d'écrire '{fname}' : le fichier est peut-être ouvert dans Excel. "
                f"Fermez-le puis relancez la sauvegarde."
            )

    def style_header(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='2E4057')
            cell.alignment = Alignment(horizontal='center')

    # ── 0. PORTEFEUILLES RÉELS ─────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portefeuilles'
    headers_pf = ['Nom', 'Description', 'Type', 'Courtier', 'Devise', 'Date ouverture', 'Couleur']
    ws.append(headers_pf)
    style_header(ws, len(headers_pf))
    for pf in file_storage.get_all_portfolios():
        ws.append([
            pf.get('name', ''),
            pf.get('description', '') or '',
            pf.get('type_compte', '') or '',
            pf.get('courtier', '') or '',
            pf.get('devise', 'EUR') or 'EUR',
            pf.get('date_ouverture', '') or '',
            pf.get('couleur', '') or '',
        ])
    pf_path = os.path.join(SAUVEGARDE_DIR, f'portefeuilles_{today}.xlsx')
    safe_save(wb, pf_path)
    fichiers_crees.append(f'portefeuilles_{today}.xlsx')

    # ── 1. CASH ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cash'
    headers_cash = ['Portefeuille', 'Banque', 'Montant', 'Devise', 'Date', 'Commentaire']
    ws.append(headers_cash)
    style_header(ws, len(headers_cash))
    for c in file_storage.get_all_cash():
        portfolio = file_storage.get_portfolio_by_id(c['portfolio_id'])
        ws.append([
            portfolio['name'] if portfolio else str(c['portfolio_id']),
            c.get('banque', ''),
            float(c['montant']) if c.get('montant') is not None else '',
            c.get('devise', 'EUR'),
            c.get('date', ''),
            c.get('commentaire', '') or '',
        ])
    cash_path = os.path.join(SAUVEGARDE_DIR, f'lastcash_{today}.xlsx')
    safe_save(wb, cash_path)
    fichiers_crees.append(f'lastcash_{today}.xlsx')

    # ── 2. PORTEFEUILLES CIBLES ───────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portefeuilles cibles'
    headers_tp = ['Portefeuille', 'ISIN', 'Titre', 'Ratio']
    ws.append(headers_tp)
    style_header(ws, len(headers_tp))
    for tp in file_storage.get_all_target_portfolios():
        for item in (tp.get('items') or []):
            sig = file_storage.get_signaletique_by_id(item.get('signaletique'))
            ws.append([
                tp['name'],
                sig['isin'] if sig else '',
                sig['titre'] if sig else '',
                item.get('ratio', 0),
            ])
    tp_path = os.path.join(SAUVEGARDE_DIR, f'portefeuillecible_{today}.xlsx')
    safe_save(wb, tp_path)
    fichiers_crees.append(f'portefeuillecible_{today}.xlsx')

    # ── 3. TRANSACTIONS ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    headers_tx = ['Date', 'Type', 'Isin', 'Quantité', 'Prix unitaire', 'Devise', 'Portefeuille']
    ws.append(headers_tx)
    style_header(ws, len(headers_tx))
    for tx in file_storage.get_all_transactions():
        sig = file_storage.get_signaletique_for_transaction(tx)
        portfolio = file_storage.get_portfolio_by_id(tx.get('portfolio_id'))
        ws.append([
            tx.get('date', ''),
            tx.get('type_operation', ''),
            sig['isin'] if sig else '',
            float(tx['quantite']) if tx.get('quantite') is not None else '',
            float(tx['prix_unitaire']) if tx.get('prix_unitaire') is not None else '',
            tx.get('devise', 'EUR'),
            portfolio['name'] if portfolio else str(tx.get('portfolio_id', '')),
        ])
    tx_path = os.path.join(SAUVEGARDE_DIR, f'transactions_{today}.xlsx')
    safe_save(wb, tx_path)
    fichiers_crees.append(f'transactions_{today}.xlsx')

    # ── 4. SIGNALÉTIQUE ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Signaletique'
    # Collecter tous les champs supplementaires
    sigs = file_storage.get_all_signaletiques()
    LAST_COLS = ['Frequence Coupon']  # colonnes toujours présentes en dernière position
    extra_fields = []
    seen = set(LAST_COLS)  # exclure des colonnes dynamiques pour les placer à la fin
    for s in sigs:
        for k in (s.get('donnees_supplementaires') or {}):
            if k not in seen:
                extra_fields.append(k)
                seen.add(k)
    extra_fields += LAST_COLS  # ajouter en dernier, qu'il y ait des données ou non
    headers_sig = ['Isin', 'Nom', "Classe d'actifs", "Type d'instr"] + extra_fields
    ws.append(headers_sig)
    style_header(ws, len(headers_sig))
    for s in sigs:
        ds = s.get('donnees_supplementaires') or {}
        row = [
            s.get('isin', '') or '',
            s.get('titre', ''),
            s.get('categorie_text', '') or '',
            s.get('statut', '') or '',
        ] + [ds.get(f, '') for f in extra_fields]
        ws.append(row)
    sig_path = os.path.join(SAUVEGARDE_DIR, f'signaletique_{today}.xlsx')
    safe_save(wb, sig_path)
    fichiers_crees.append(f'signaletique_{today}.xlsx')

    # ── 5. UTILISATEURS ───────────────────────────────────────────────────────────────────────
    from django.contrib.auth.models import User as DjangoUser
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Utilisateurs'
    headers_users = ['username', 'email', 'first_name', 'last_name', 'is_staff', 'is_superuser', 'is_active', 'password_hash']
    ws.append(headers_users)
    style_header(ws, len(headers_users))
    for u in DjangoUser.objects.all().order_by('id'):
        ws.append([
            u.username,
            u.email or '',
            u.first_name or '',
            u.last_name or '',
            u.is_staff,
            u.is_superuser,
            u.is_active,
            u.password,  # hash Django (pbkdf2_sha256$...)
        ])
    users_path = os.path.join(SAUVEGARDE_DIR, f'utilisateurs_{today}.xlsx')
    safe_save(wb, users_path)
    fichiers_crees.append(f'utilisateurs_{today}.xlsx')
    nb_users = DjangoUser.objects.count()

    # ── 6. PRIX HISTORIQUE TITRES ─────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Prix Historique'
    headers_prix = ['ISIN', 'Nom', 'Date', 'Cours', 'Devise', 'Source', 'Symbole']
    ws.append(headers_prix)
    style_header(ws, len(headers_prix))
    prix_titres = file_storage.get_prix_historique_titres()
    nb_prix = 0
    for isin, titre_data in sorted(prix_titres.items()):
        nom = titre_data.get('nom', '')
        for entry in (titre_data.get('historique') or []):
            ws.append([
                isin,
                nom,
                entry.get('date', ''),
                float(entry['cours']) if entry.get('cours') is not None else '',
                entry.get('devise', ''),
                entry.get('source', ''),
                entry.get('symbole', '') or '',
            ])
            nb_prix += 1
    prix_path = os.path.join(SAUVEGARDE_DIR, f'prix_historique_{today}.xlsx')
    safe_save(wb, prix_path)
    fichiers_crees.append(f'prix_historique_{today}.xlsx')

    return Response({
        'success': True,
        'date': today,
        'repertoire': SAUVEGARDE_DIR,
        'fichiers': fichiers_crees,
        'stats': {
            'portefeuilles': len(file_storage.get_all_portfolios()),
            'cash': len(file_storage.get_all_cash()),
            'portefeuilles_cibles': len(file_storage.get_all_target_portfolios()),
            'transactions': len(file_storage.get_all_transactions()),
            'signaletiques': len(sigs),
            'utilisateurs': nb_users,
            'prix_historique': nb_prix,
        }
    })


@api_view(['POST'])
def restore_all_data(request):
    """Restaure toutes les données depuis les fichiers les plus récents de /app/sauvegarde/.
    Ordre : signalétique → transactions → cash → portefeuilles cibles → utilisateurs.
    Réservé aux superusers.
    """
    if not request.user.is_superuser:
        return Response({'error': 'Réservé aux superusers'}, status=status.HTTP_403_FORBIDDEN)

    import glob
    from django.contrib.auth.models import User as DjangoUser

    SAUVEGARDE_DIR = '/app/startfiles'
    results = {}

    def latest_file(prefix):
        files = sorted(glob.glob(os.path.join(SAUVEGARDE_DIR, f'{prefix}_*.xlsx')))
        return files[-1] if files else None

    def sanitize_date(v):
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        return v

    # ── 0. PORTEFEUILLES RÉELS ────────────────────────────────────────────────
    # Priorité 1 : snapshot JSON (préserve IDs + tous les champs)
    # Priorité 2 : fichier Excel (reconstruit par nom, IDs recalculés)
    snapshots = file_storage.list_json_snapshots(SAUVEGARDE_DIR)
    pf_json_restored = False
    if snapshots:
        latest_snap = snapshots[-1]
        snap_pf_file = os.path.join(SAUVEGARDE_DIR, latest_snap, 'portfolios.json')
        if os.path.exists(snap_pf_file):
            try:
                pf_data = file_storage._read_json_file(snap_pf_file, [])
                file_storage._write_json_file(file_storage.PORTFOLIOS_FILE, pf_data)
                results['portefeuilles'] = {
                    'succes': len(pf_data), 'erreurs': 0,
                    'source': f'snapshot JSON ({latest_snap})',
                    'note': 'IDs et tous les champs préservés',
                }
                pf_json_restored = True
            except Exception as e:
                results['portefeuilles'] = {'error': f'Erreur snapshot JSON : {e}'}

    # Dictionnaire de référence nom → métadonnées, utilisé comme fallback
    # dans toutes les sections qui créent des portefeuilles à la volée.
    portfolio_meta_ref: dict = {}

    if not pf_json_restored:
        pf_xlsx_path = latest_file('portefeuilles')
        if pf_xlsx_path:
            try:
                wb = openpyxl.load_workbook(pf_xlsx_path)
                ws = wb.active
                headers_pf = [str(c.value or '').strip() for c in ws[1]]
                succes_pf = 0
                erreurs_pf = []
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(c is not None and str(c).strip() != '' for c in row):
                        continue
                    row_data = {h: v for h, v in zip(headers_pf, row)}
                    nom = str(row_data.get('Nom') or '').strip()
                    if not nom:
                        erreurs_pf.append(f"Ligne {row_idx}: nom manquant")
                        continue
                    pf_fields = {
                        'description':    str(row_data.get('Description') or '') or None,
                        'type_compte':    str(row_data.get('Type') or '') or None,
                        'courtier':       str(row_data.get('Courtier') or '') or None,
                        'devise':         str(row_data.get('Devise') or 'EUR'),
                        'date_ouverture': str(row_data.get('Date ouverture') or '') or None,
                        'couleur':        str(row_data.get('Couleur') or '') or None,
                    }
                    # Mémoriser pour les créations à la volée dans les sections suivantes
                    portfolio_meta_ref[nom] = pf_fields
                    existing = file_storage.get_portfolio_by_name(nom)
                    if existing:
                        file_storage.update_portfolio(existing['id'], pf_fields)
                    else:
                        file_storage.create_portfolio(name=nom, **pf_fields)
                    succes_pf += 1
                results['portefeuilles'] = {
                    'succes': succes_pf, 'erreurs': len(erreurs_pf),
                    'source': f'Excel ({os.path.basename(pf_xlsx_path)})',
                    'note': 'IDs recalculés — snapshot JSON non disponible',
                }
            except Exception as e:
                results['portefeuilles'] = {'succes': 0, 'erreurs': 1, 'fichier': os.path.basename(pf_xlsx_path), 'error': str(e)}
        else:
            results['portefeuilles'] = {'info': 'Aucune source disponible (ni snapshot JSON ni portefeuilles_*.xlsx)'}

    # ── 1. SIGNALÉTIQUE ──────────────────────────────────────────────────────
    sig_path = latest_file('signaletique')
    if sig_path:
        try:
            wb = openpyxl.load_workbook(sig_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            succes = 0
            erreurs_list = []

            def _sanitize_sig(value):
                if isinstance(value, (datetime, date)):
                    return value.isoformat()
                return value

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(c is not None and str(c).strip() != '' for c in row):
                    continue
                row_data = {h: _sanitize_sig(v) for h, v in zip(headers, row)}
                isin_raw = row_data.get('Isin') or row_data.get('ISIN') or row_data.get('isin')
                isin_value = str(isin_raw).strip().upper() if isin_raw else None
                if not isin_value:
                    isin_value = None
                if isin_value and len(isin_value) != 12:
                    erreurs_list.append(f"Ligne {row_idx}: ISIN invalide ({isin_value})")
                    continue
                code = f"SIG_{isin_value}" if isin_value else f"AUTO_{row_idx}"
                titre = row_data.get('Nom') or ''
                categorie_text = str(row_data.get("Classe d'actifs") or '').strip().capitalize() or None
                statut_data = str(row_data.get("Type d'instr") or '')
                sig_dict = file_storage.upsert_signaletique(
                    code=str(code), isin=isin_value, titre=str(titre)[:500],
                    categorie_text=categorie_text,
                    statut=statut_data[:100] if statut_data else None,
                    donnees_supplementaires=row_data,
                )
                categorie_instance = None
                if categorie_text:
                    categorie_instance, _ = AssetCategory.objects.get_or_create(
                        name=categorie_text,
                        defaults={'description': 'Catégorie créée automatiquement lors de la restauration'},
                    )
                if not isin_value and not str(titre).strip():
                    continue
                db_defaults = {
                    'code': str(code), 'isin': isin_value, 'titre': str(titre)[:500],
                    'categorie': categorie_instance, 'categorie_text': categorie_text,
                    'statut': statut_data[:100] if statut_data else None,
                    'donnees_supplementaires': row_data,
                }
                if isin_value:
                    Signaletique.objects.update_or_create(isin=isin_value, defaults=db_defaults)
                else:
                    Signaletique.objects.update_or_create(code=str(code), defaults=db_defaults)
                succes += 1

            results['signaletique'] = {'succes': succes, 'erreurs': len(erreurs_list), 'fichier': os.path.basename(sig_path)}
        except Exception as e:
            results['signaletique'] = {'succes': 0, 'erreurs': 1, 'fichier': os.path.basename(sig_path), 'error': str(e)}
    else:
        results['signaletique'] = {'error': 'Aucun fichier signaletique_*.xlsx trouvé dans Sauvegarde/'}

    # ── 2. TRANSACTIONS ───────────────────────────────────────────────────────
    tx_path = latest_file('transactions')
    if tx_path:
        try:
            wb = openpyxl.load_workbook(tx_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            succes = 0
            doublons = 0
            erreurs_list = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(c is not None and str(c).strip() != '' for c in row):
                    continue
                row_data = {h: sanitize_date(v) for h, v in zip(headers, row)}
                date_tx = row_data.get('Date')
                type_op = str(row_data.get('Type') or row_data.get('Sens') or '').strip().upper()
                isin = row_data.get('Isin') or row_data.get('ISIN') or row_data.get('isin')
                if isin:
                    isin = str(isin).strip().upper()
                quantite = row_data.get('Quantité') or row_data.get('quantite') or row_data.get('Quantite')
                prix = row_data.get('Prix unitaire') or row_data.get('prix unitaire') or row_data.get('Prix Unitaire')
                devise = str(row_data.get('Devise') or 'EUR')
                nom_portfolio = str(row_data.get('Portefeuille') or 'Défaut')

                if not all([type_op, isin, quantite, prix]):
                    erreurs_list.append(f"Ligne {row_idx}: champs obligatoires manquants")
                    continue
                if len(isin) != 12:
                    erreurs_list.append(f"Ligne {row_idx}: ISIN invalide ({isin})")
                    continue
                if type_op not in ['ACHAT', 'VENTE']:
                    type_op = 'ACHAT' if 'achat' in type_op.lower() else 'VENTE'

                portfolio, _ = file_storage.get_or_create_portfolio(
                    nom_portfolio,
                    defaults=portfolio_meta_ref.get(nom_portfolio, {'description': f'Portefeuille {nom_portfolio}'})
                )
                sig_dict = file_storage.get_signaletique_by_isin(isin)
                if not sig_dict:
                    sig_dict = file_storage.upsert_signaletique(
                        code=f'TEMP_{isin}', isin=isin, titre=f'[À compléter] {isin}'
                    )
                    Signaletique.objects.get_or_create(
                        isin=isin, defaults={'code': f'TEMP_{isin}', 'titre': f'[À compléter] {isin}'}
                    )

                q = Decimal(str(quantite))
                p = Decimal(str(prix))
                date_str = date_tx.isoformat() if hasattr(date_tx, 'isoformat') else str(date_tx)

                if file_storage.transaction_exists(portfolio['id'], sig_dict['id'], date_str, type_op, q, p, isin=isin):
                    doublons += 1
                    continue

                file_storage.create_transaction(portfolio['id'], sig_dict['id'], date_str, type_op, q, p, devise, isin=isin)
                succes += 1

            results['transactions'] = {
                'succes': succes, 'doublons': doublons,
                'erreurs': len(erreurs_list), 'fichier': os.path.basename(tx_path),
            }
        except Exception as e:
            results['transactions'] = {'succes': 0, 'erreurs': 1, 'fichier': os.path.basename(tx_path), 'error': str(e)}
    else:
        results['transactions'] = {'error': 'Aucun fichier transactions_*.xlsx trouvé dans Sauvegarde/'}

    # ── 3. CASH ───────────────────────────────────────────────────────────────
    cash_path = latest_file('lastcash')
    if cash_path:
        try:
            wb = openpyxl.load_workbook(cash_path)
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in ws[1]]
            succes = 0
            erreurs_list = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(c is not None and str(c).strip() != '' for c in row):
                    continue
                row_data = {h: v for h, v in zip(headers, row)}
                nom_portfolio = str(row_data.get('Portefeuille') or '').strip()
                banque = str(row_data.get('Banque') or '').strip()
                montant_raw = row_data.get('Montant')
                devise = str(row_data.get('Devise') or 'EUR')
                date_raw = row_data.get('Date')
                commentaire = str(row_data.get('Commentaire') or '').strip() or None

                if not all([nom_portfolio, banque, montant_raw, date_raw]):
                    erreurs_list.append(f"Ligne {row_idx}: champs manquants")
                    continue

                montant = Decimal(str(montant_raw))
                if isinstance(date_raw, datetime):
                    date_obj = date_raw.date()
                elif isinstance(date_raw, date):
                    date_obj = date_raw
                else:
                    date_obj = datetime.fromisoformat(str(date_raw)).date()

                portfolio, _ = file_storage.get_or_create_portfolio(
                    nom_portfolio,
                    defaults=portfolio_meta_ref.get(nom_portfolio, {'description': f'Portefeuille {nom_portfolio}'})
                )
                file_storage.create_cash(portfolio['id'], banque, montant, devise, date_obj, commentaire)
                succes += 1

            results['cash'] = {'succes': succes, 'erreurs': len(erreurs_list), 'fichier': os.path.basename(cash_path)}
        except Exception as e:
            results['cash'] = {'succes': 0, 'erreurs': 1, 'fichier': os.path.basename(cash_path), 'error': str(e)}
    else:
        results['cash'] = {'error': 'Aucun fichier lastcash_*.xlsx trouvé dans Sauvegarde/'}

    # ── 4. PORTEFEUILLES CIBLES ───────────────────────────────────────────────
    tp_path = latest_file('portefeuillecible')
    if tp_path:
        try:
            wb = openpyxl.load_workbook(tp_path)
            ws = wb.active
            headers = [str(cell.value or '').strip() for cell in ws[1]]
            portfolios_data = {}
            succes = 0
            erreurs_list = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(c is not None and str(c).strip() != '' for c in row):
                    continue
                row_data = {h: v for h, v in zip(headers, row)}
                nom_p = str(row_data.get('Portefeuille') or '').strip()
                titre_raw = str(row_data.get('ISIN') or row_data.get('Isin') or row_data.get('Titre') or '').strip()
                ratio_raw = row_data.get('Ratio')

                if not all([nom_p, titre_raw, ratio_raw]):
                    erreurs_list.append(f"Ligne {row_idx}: champs manquants")
                    continue

                sig = (
                    file_storage.get_signaletique_by_isin(titre_raw) or
                    file_storage.get_signaletique_by_code(titre_raw) or
                    file_storage.search_signaletique_by_titre(titre_raw)
                )
                if not sig:
                    erreurs_list.append(f"Ligne {row_idx}: ISIN/titre introuvable ({titre_raw})")
                    continue

                if nom_p not in portfolios_data:
                    portfolios_data[nom_p] = []
                portfolios_data[nom_p].append({'signaletique': sig['id'], 'ratio': float(str(ratio_raw))})

            for nom, items in portfolios_data.items():
                existing = file_storage.get_target_portfolio_by_name(nom)
                if existing:
                    file_storage.update_target_portfolio(existing['id'], name=nom, items=items)
                else:
                    file_storage.create_target_portfolio(nom, items)
                succes += 1

            results['portefeuilles_cibles'] = {
                'succes': succes, 'erreurs': len(erreurs_list), 'fichier': os.path.basename(tp_path),
            }
        except Exception as e:
            results['portefeuilles_cibles'] = {'succes': 0, 'erreurs': 1, 'fichier': os.path.basename(tp_path), 'error': str(e)}
    else:
        results['portefeuilles_cibles'] = {'error': 'Aucun fichier portefeuillecible_*.xlsx trouvé dans Sauvegarde/'}

    # ── 5. UTILISATEURS ───────────────────────────────────────────────────────
    users_path = latest_file('utilisateurs')
    if users_path:
        try:
            wb = openpyxl.load_workbook(users_path)
            ws = wb.active
            headers = [str(c.value or '').strip() for c in ws[1]]
            crees = 0
            mis_a_jour = 0
            erreurs_list = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {h: v for h, v in zip(headers, row)}
                username = str(row_data.get('username') or '').strip()
                if not username:
                    continue
                password_hash = str(row_data.get('password_hash') or '').strip()
                if not password_hash:
                    erreurs_list.append(f"Ligne {row_idx}: password_hash manquant pour {username}")
                    continue
                try:
                    user, created = DjangoUser.objects.get_or_create(
                        username=username,
                        defaults={
                            'email': str(row_data.get('email') or ''),
                            'first_name': str(row_data.get('first_name') or ''),
                            'last_name': str(row_data.get('last_name') or ''),
                            'is_staff': bool(row_data.get('is_staff')),
                            'is_superuser': bool(row_data.get('is_superuser')),
                            'is_active': row_data.get('is_active') != False,
                        },
                    )
                    user.password = password_hash
                    if not created:
                        user.email = str(row_data.get('email') or '')
                        user.first_name = str(row_data.get('first_name') or '')
                        user.last_name = str(row_data.get('last_name') or '')
                        user.is_staff = bool(row_data.get('is_staff'))
                        user.is_superuser = bool(row_data.get('is_superuser'))
                        user.is_active = row_data.get('is_active') != False
                    user.save()
                    if created:
                        crees += 1
                    else:
                        mis_a_jour += 1
                except Exception as e:
                    erreurs_list.append(f"Ligne {row_idx}: {str(e)}")

            results['utilisateurs'] = {
                'crees': crees, 'mis_a_jour': mis_a_jour,
                'erreurs': len(erreurs_list), 'fichier': os.path.basename(users_path),
            }
        except Exception as e:
            results['utilisateurs'] = {'succes': 0, 'erreurs': 1, 'fichier': os.path.basename(users_path), 'error': str(e)}
    else:
        results['utilisateurs'] = {'error': 'Aucun fichier utilisateurs_*.xlsx trouvé dans Sauvegarde/'}

    return Response({'success': True, 'details': results})


@api_view(['POST'])
def clear_all_data(request):
    """Supprime tous les fichiers JSON de /app/data/ et tous les fichiers Excel de /app/sauvegarde/.
    Ne supprime aucun fichier non-JSON / non-xlsx (ex : README.md).
    Réservé aux superusers.
    """
    if not request.user.is_superuser:
        return Response({'error': 'Réservé aux superusers'}, status=status.HTTP_403_FORBIDDEN)

    import glob

    deleted = []
    for f in glob.glob('/app/data/*.json'):
        os.remove(f)
        deleted.append(os.path.basename(f))
    for f in glob.glob('/app/sauvegarde/*.xlsx'):
        os.remove(f)
        deleted.append(os.path.basename(f))

    return Response({'success': True, 'supprime': deleted, 'total': len(deleted)})


@api_view(['GET', 'PUT', 'DELETE'])
def signaletique_detail(request, pk):
    """Récupérer, modifier ou supprimer une signalétique"""
    sig_dict = file_storage.get_signaletique_by_id(int(pk))
    if not sig_dict:
        return Response({'error': 'Signalétique non trouvée'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response(sig_dict)

    elif request.method == 'PUT':
        allowed = ['titre', 'description', 'categorie_text', 'statut', 'donnees_supplementaires']
        update_data = {k: v for k, v in request.data.items() if k in allowed}
        updated = file_storage.update_signaletique(int(pk), update_data)
        # Sync DB
        try:
            db_sig = Signaletique.objects.get(pk=pk)
            for k, v in update_data.items():
                setattr(db_sig, k, v)
            db_sig.save()
        except Signaletique.DoesNotExist:
            pass
        return Response(updated)

    elif request.method == 'DELETE':
        return Response(
            {'error': 'La suppression des signalétiques est désactivée pour préserver l\'intégrité des données'},
            status=status.HTTP_403_FORBIDDEN
        )


@api_view(['GET'])
def import_logs(request):
    """Liste des logs d'import"""
    logs = ImportLog.objects.all()[:20]  # 20 derniers logs
    serializer = ImportLogSerializer(logs, many=True)
    return Response(serializer.data)


@api_view(['GET', 'POST'])
def list_target_portfolios(request):
    if request.method == 'GET':
        return Response(file_storage.get_all_target_portfolios())

    elif request.method == 'POST':
        name = request.data.get('name', '').strip()
        items = request.data.get('items', [])
        if not name:
            return Response({'error': 'Le nom est requis'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            portfolio = file_storage.create_target_portfolio(name, items)
            return Response(portfolio, status=status.HTTP_201_CREATED)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
def target_portfolio_detail(request, pk):
    portfolio = file_storage.get_target_portfolio_by_id(int(pk))
    if not portfolio:
        return Response({'error': 'Portefeuille cible non trouvé'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        return Response(portfolio)

    elif request.method == 'PUT':
        name = request.data.get('name', '').strip() or None
        items = request.data.get('items', None)
        try:
            updated = file_storage.update_target_portfolio(int(pk), name=name, items=items)
            if updated:
                return Response(updated)
            return Response({'error': 'Portefeuille cible non trouvé'}, status=status.HTTP_404_NOT_FOUND)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        if file_storage.delete_target_portfolio(int(pk)):
            return Response({'success': True, 'message': 'Portefeuille cible supprimé'}, status=status.HTTP_204_NO_CONTENT)
        return Response({'error': 'Portefeuille cible non trouvé'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET', 'POST'])
def list_real_portfolios(request):
    """Liste ou crée des portefeuilles réels"""
    if request.method == 'GET':
        portfolios = file_storage.get_all_portfolios()
        return Response(portfolios)
    
    elif request.method == 'POST':
        name = request.data.get('name')
        if not name:
            return Response(
                {'error': 'Le nom du portefeuille est requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            portfolio = file_storage.create_portfolio(
                name=name,
                description=request.data.get('description'),
                type_compte=request.data.get('type_compte'),
                courtier=request.data.get('courtier'),
                devise=request.data.get('devise', 'EUR'),
                date_ouverture=request.data.get('date_ouverture'),
                couleur=request.data.get('couleur'),
            )
            return Response(portfolio, status=status.HTTP_201_CREATED)
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


@api_view(['GET', 'PUT', 'DELETE'])
def real_portfolio_detail(request, pk):
    """Récupère, modifie ou supprime un portefeuille réel"""
    portfolio = file_storage.get_portfolio_by_id(int(pk))
    
    if not portfolio:
        return Response(
            {'error': 'Portefeuille non trouve'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if request.method == 'GET':
        return Response(portfolio)
    
    elif request.method == 'PUT':
        try:
            updated_portfolio = file_storage.update_portfolio(int(pk), request.data)
            if updated_portfolio:
                return Response(updated_portfolio)
            return Response(
                {'error': 'Portefeuille non trouve'},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    elif request.method == 'DELETE':
        if file_storage.delete_portfolio(int(pk)):
            return Response(
                {'success': True, 'message': 'Portefeuille supprime avec toutes ses transactions et cash'},
                status=status.HTTP_204_NO_CONTENT
            )
        return Response(
            {'error': 'Portefeuille non trouve'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET', 'POST'])
def list_cash(request):
    """Liste ou crée des entrées de cash"""
    if request.method == 'GET':
        # Optionnel: filtrer par portfolio si fourni dans les params
        portfolio_id = request.query_params.get('portfolio_id', None)
        if portfolio_id:
            cash_entries = file_storage.get_cash_by_portfolio(int(portfolio_id))
        else:
            cash_entries = file_storage.get_all_cash()
        return Response(cash_entries)
    
    elif request.method == 'POST':
        portfolio_id = request.data.get('portfolio_id') or request.data.get('portfolio')
        banque = request.data.get('banque')
        montant = request.data.get('montant')
        devise = request.data.get('devise', 'EUR')
        date_str = request.data.get('date')
        commentaire = request.data.get('commentaire')
        
        # Validation
        if not all([portfolio_id, banque, montant, date_str]):
            return Response(
                {'error': 'Les champs portfolio_id, banque, montant et date sont requis'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Vérifier que le portefeuille existe
        portfolio = file_storage.get_portfolio_by_id(int(portfolio_id))
        if not portfolio:
            return Response(
                {'error': 'Portefeuille non trouvé'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Convertir la date
        try:
            if isinstance(date_str, str):
                date_obj = datetime.fromisoformat(date_str).date()
            else:
                date_obj = date_str
        except (ValueError, AttributeError):
            return Response(
                {'error': 'Format de date invalide. Utilisez le format ISO (YYYY-MM-DD)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            cash = file_storage.create_cash(
                int(portfolio_id),
                banque,
                Decimal(str(montant)),
                devise,
                date_obj,
                commentaire
            )
            return Response(cash, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response(
                {'error': f'Erreur lors de la création: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


@api_view(['GET', 'PUT', 'DELETE'])
def cash_detail(request, pk):
    """Récupère, modifie ou supprime une entrée de cash"""
    cash = file_storage.get_cash_by_id(int(pk))
    
    if not cash:
        return Response(
            {'error': 'Entree de cash non trouvee'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if request.method == 'GET':
        return Response(cash)
    
    elif request.method == 'PUT':
        # Préparer les données à mettre à jour
        update_data = {}
        if 'banque' in request.data:
            update_data['banque'] = request.data['banque']
        if 'montant' in request.data:
            update_data['montant'] = str(Decimal(str(request.data['montant'])))
        if 'devise' in request.data:
            update_data['devise'] = request.data['devise']
        if 'date' in request.data:
            date_str = request.data['date']
            if isinstance(date_str, str):
                update_data['date'] = datetime.fromisoformat(date_str).date().isoformat()
            else:
                update_data['date'] = date_str
        if 'commentaire' in request.data:
            update_data['commentaire'] = request.data['commentaire']
        
        updated_cash = file_storage.update_cash(int(pk), update_data)
        if updated_cash:
            return Response(updated_cash)
        return Response(
            {'error': 'Entree de cash non trouvee'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    elif request.method == 'DELETE':
        if file_storage.delete_cash(int(pk)):
            return Response(
                {'success': True, 'message': 'Entree de cash supprimee'},
                status=status.HTTP_204_NO_CONTENT
            )
        return Response(
            {'error': 'Entree de cash non trouvee'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_cash(request):
    """Import des positions cash depuis un fichier Excel.
    Colonnes attendues: Portefeuille, Banque, Montant, Devise, Date, Commentaire
    """
    if 'file' not in request.FILES:
        return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    if not file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'Format non supporté. Utilisez .xlsx ou .xls'}, status=status.HTTP_400_BAD_REQUEST)

    import_log = ImportLog.objects.create(
        type_import='cash',
        nom_fichier=file.name,
        statut='en_cours'
    )

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]

        nombre_succes = 0
        nombre_erreurs = 0
        erreurs = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() != '' for cell in row):
                continue

            row_data = {header: value for header, value in zip(headers, row)}

            nom_portfolio = (
                str(row_data.get('Portefeuille') or row_data.get('portefeuille') or '').strip()
            )
            banque = str(row_data.get('Banque') or row_data.get('banque') or '').strip()
            montant_raw = row_data.get('Montant') or row_data.get('montant')
            devise = str(row_data.get('Devise') or row_data.get('devise') or 'EUR').strip()
            date_raw = row_data.get('Date') or row_data.get('date')
            commentaire = str(row_data.get('Commentaire') or row_data.get('commentaire') or '').strip() or None

            if not all([nom_portfolio, banque, montant_raw, date_raw]):
                missing = []
                if not nom_portfolio: missing.append('Portefeuille')
                if not banque: missing.append('Banque')
                if not montant_raw: missing.append('Montant')
                if not date_raw: missing.append('Date')
                erreurs.append({'ligne': row_idx, 'type': 'erreur', 'erreur': f'Champs manquants: {", ".join(missing)}'})
                nombre_erreurs += 1
                continue

            try:
                montant_decimal = Decimal(str(montant_raw))
            except Exception:
                erreurs.append({'ligne': row_idx, 'type': 'erreur', 'erreur': f'Montant invalide: {montant_raw}'})
                nombre_erreurs += 1
                continue

            if isinstance(date_raw, (datetime, date)):
                date_obj = date_raw.date() if isinstance(date_raw, datetime) else date_raw
            else:
                try:
                    date_obj = datetime.fromisoformat(str(date_raw)).date()
                except ValueError:
                    erreurs.append({'ligne': row_idx, 'type': 'erreur', 'erreur': f'Date invalide: {date_raw}'})
                    nombre_erreurs += 1
                    continue

            portfolio, _ = file_storage.get_or_create_portfolio(
                nom_portfolio,
                defaults={'description': f'Portefeuille {nom_portfolio}'}
            )

            file_storage.create_cash(portfolio['id'], banque, montant_decimal, devise, date_obj, commentaire)
            nombre_succes += 1

        import_log.nombre_lignes = ws.max_row - 1
        import_log.nombre_succes = nombre_succes
        import_log.nombre_erreurs = nombre_erreurs
        import_log.erreurs = erreurs if erreurs else None
        import_log.statut = 'termine' if nombre_erreurs == 0 else 'termine_avec_erreurs'
        import_log.save()

        return Response({
            'success': True,
            'message': 'Import cash terminé',
            'details': {
                'fichier': file.name,
                'lignes_totales': ws.max_row - 1,
                'succes': nombre_succes,
                'erreurs': nombre_erreurs,
                'liste_erreurs': erreurs[:10] if erreurs else []
            }
        })

    except Exception as e:
        import_log.statut = 'erreur'
        import_log.save()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_target_portfolio(request):
    """Import de portefeuilles cibles depuis un fichier Excel.
    Colonnes attendues: Portefeuille, Titre (nom/ISIN/code du titre), Ratio
    Plusieurs lignes avec le même nom de portefeuille → un seul portefeuille.
    """
    if 'file' not in request.FILES:
        return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    if not file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'Format non supporté. Utilisez .xlsx ou .xls'}, status=status.HTTP_400_BAD_REQUEST)

    import_log = ImportLog.objects.create(
        type_import='target_portfolio',
        nom_fichier=file.name,
        statut='en_cours'
    )

    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]

        nombre_succes = 0
        nombre_erreurs = 0
        erreurs = []

        # Regrouper les lignes par nom de portefeuille
        portfolios_data = {}  # name → list of items

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() != '' for cell in row):
                continue

            row_data = {header: value for header, value in zip(headers, row)}

            nom_portfolio = str(row_data.get('Portefeuille') or row_data.get('portefeuille') or '').strip()
            titre_raw = str(row_data.get('Titre') or row_data.get('titre') or row_data.get('ISIN') or row_data.get('isin') or '').strip()
            ratio_raw = row_data.get('Ratio') or row_data.get('ratio') or row_data.get('%') or row_data.get('Poids')

            if not all([nom_portfolio, titre_raw, ratio_raw]):
                missing = []
                if not nom_portfolio: missing.append('Portefeuille')
                if not titre_raw: missing.append('Titre')
                if not ratio_raw: missing.append('Ratio')
                erreurs.append({'ligne': row_idx, 'type': 'erreur', 'erreur': f'Champs manquants: {", ".join(missing)}'})
                nombre_erreurs += 1
                continue

            try:
                ratio = float(str(ratio_raw))
            except ValueError:
                erreurs.append({'ligne': row_idx, 'type': 'erreur', 'erreur': f'Ratio invalide: {ratio_raw}'})
                nombre_erreurs += 1
                continue

            # Résoudre le titre : chercher par ISIN, puis code, puis titre
            sig = (
                file_storage.get_signaletique_by_isin(titre_raw) or
                file_storage.get_signaletique_by_code(titre_raw) or
                file_storage.search_signaletique_by_titre(titre_raw)
            )

            if not sig:
                erreurs.append({'ligne': row_idx, 'type': 'erreur', 'erreur': f'Titre introuvable: {titre_raw}'})
                nombre_erreurs += 1
                continue

            if nom_portfolio not in portfolios_data:
                portfolios_data[nom_portfolio] = []
            portfolios_data[nom_portfolio].append({'signaletique': sig['id'], 'ratio': ratio})

        # Créer ou mettre à jour les portefeuilles cibles via file_storage
        for nom, items in portfolios_data.items():
            try:
                existing = file_storage.get_target_portfolio_by_name(nom)
                if existing:
                    file_storage.update_target_portfolio(existing['id'], name=nom, items=items)
                else:
                    file_storage.create_target_portfolio(nom, items)
                nombre_succes += 1
            except ValueError as e:
                erreurs.append({'ligne': 0, 'type': 'erreur', 'erreur': f'Portefeuille "{nom}": {str(e)}'})
                nombre_erreurs += 1

        import_log.nombre_lignes = ws.max_row - 1
        import_log.nombre_succes = nombre_succes
        import_log.nombre_erreurs = nombre_erreurs
        import_log.erreurs = erreurs if erreurs else None
        import_log.statut = 'termine' if nombre_erreurs == 0 else 'termine_avec_erreurs'
        import_log.save()

        return Response({
            'success': True,
            'message': 'Import portefeuilles cibles terminé',
            'details': {
                'fichier': file.name,
                'lignes_totales': ws.max_row - 1,
                'portefeuilles_crees': nombre_succes,
                'erreurs': nombre_erreurs,
                'liste_erreurs': erreurs[:10] if erreurs else []
            }
        })

    except Exception as e:
        import_log.statut = 'erreur'
        import_log.save()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_transactions(request):
    """Import des transactions d'achat/vente depuis un fichier Excel"""
    
    if 'file' not in request.FILES:
        return Response(
            {'error': 'Aucun fichier fourni'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    file = request.FILES['file']
    
    # Vérifier l'extension
    if not file.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'Format de fichier non supporté. Utilisez .xlsx ou .xls'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Créer un log d'import
    import_log = ImportLog.objects.create(
        type_import='transactions',
        nom_fichier=file.name,
        statut='en_cours'
    )
    
    try:
        # Lire le fichier Excel
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # Récupérer les en-têtes (première ligne)
        headers = [cell.value for cell in ws[1]]
        
        nombre_succes = 0
        nombre_erreurs = 0
        nombre_doublons = 0
        erreurs = []
        isins_inconnus = set()
        
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        def sanitize_value(value):
            if isinstance(value, (datetime, date)):
                return value.date() if isinstance(value, datetime) else value
            return value
        
        # Parcourir les lignes
        for row_idx, row in enumerate(rows, start=2):
            try:
                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue
                
                # Créer un dictionnaire avec les données de la ligne
                row_data = {
                    header: sanitize_value(value)
                    for header, value in zip(headers, row)
                }
                
                # Extraire les champs avec plus de flexibilité
                date_transaction = row_data.get('Date') or row_data.get('date')
                
                # Type/Sens de l'opération
                type_operation = (
                    row_data.get('Type') or 
                    row_data.get('type') or 
                    row_data.get('Sens') or 
                    row_data.get('Sens ') or 
                    row_data.get('sens') or 
                    ''
                ).strip().upper()
                
                # ISIN - normaliser en enlevant les espaces et en majuscules
                isin = row_data.get('Isin') or row_data.get('ISIN') or row_data.get('isin')
                if isin:
                    isin = str(isin).strip().upper()
                
                # Quantité
                quantite = (
                    row_data.get('quantité') or 
                    row_data.get('Quantité') or 
                    row_data.get('quantite') or
                    row_data.get('Quantite')
                )
                
                # Prix unitaire
                prix_unitaire = (
                    row_data.get('prix unitaire') or 
                    row_data.get('Prix unitaire') or 
                    row_data.get('Prix Unitaire') or
                    row_data.get('prix') or
                    row_data.get('Prix')
                )
                
                devise = row_data.get('Devise') or row_data.get('devise') or 'EUR'
                nom_portfolio = row_data.get('Portefeuille') or row_data.get('portefeuille') or 'Défaut'
                
                # Si date manquante, utiliser la date du jour
                if not date_transaction:
                    date_transaction = date.today()
                    erreurs.append({
                        'ligne': row_idx,
                        'type': 'warning',
                        'erreur': 'Date manquante, date du jour utilisée'
                    })
                
                # Validation
                if not all([type_operation, isin, quantite, prix_unitaire]):
                    missing_fields = []
                    if not type_operation: missing_fields.append('Type/Sens')
                    if not isin: missing_fields.append('ISIN')
                    if not quantite: missing_fields.append('quantité')
                    if not prix_unitaire: missing_fields.append('prix')
                    
                    erreurs.append({
                        'ligne': row_idx,
                        'type': 'erreur',
                        'erreur': f'Champs obligatoires manquants: {", ".join(missing_fields)}'
                    })
                    nombre_erreurs += 1
                    continue
                
                # Normaliser le type
                if type_operation not in ['ACHAT', 'VENTE']:
                    type_operation = 'ACHAT' if 'achat' in type_operation.lower() else 'VENTE'
                
                # Récupérer ou créer le portefeuille
                portfolio, _ = file_storage.get_or_create_portfolio(
                    nom_portfolio,
                    defaults={'description': f'Portefeuille {nom_portfolio}'}
                )
                
                # Validation stricte : l'ISIN doit avoir exactement 12 caractères
                if len(isin) != 12:
                    erreurs.append({
                        'ligne': row_idx,
                        'type': 'erreur',
                        'erreur': f"ISIN invalide (longueur {len(isin)} au lieu de 12): {isin}. Les ISINs doivent avoir exactement 12 caractères."
                    })
                    nombre_erreurs += 1
                    continue
                
                # Récupérer la signalétique - Recherche exacte uniquement
                sig_dict = file_storage.get_signaletique_by_isin(isin)
                if sig_dict:
                    signaletique_id = sig_dict['id']
                else:
                    # ISIN inconnu - créer dans JSON + DB
                    isins_inconnus.add(isin)
                    sig_dict = file_storage.upsert_signaletique(
                        code=f'TEMP_{isin}', isin=isin,
                        titre=f'[À compléter] {isin}'
                    )
                    signaletique_id = sig_dict['id']
                    # Double-write DB
                    db_sig, _ = Signaletique.objects.get_or_create(
                        isin=isin,
                        defaults={'code': f'TEMP_{isin}', 'titre': f'[À compléter] {isin}'}
                    )
                    erreurs.append({
                        'ligne': row_idx,
                        'type': 'warning',
                        'erreur': f"L'ISIN {isin} est inconnu du système portfolio"
                    })
                
                # Vérifier si la transaction existe déjà (doublon)
                quantite_decimal = Decimal(str(quantite))
                prix_unitaire_decimal = Decimal(str(prix_unitaire))
                
                existing_transaction = file_storage.transaction_exists(
                    portfolio['id'],
                    signaletique_id,
                    date_transaction.isoformat() if hasattr(date_transaction, 'isoformat') else date_transaction,
                    type_operation,
                    quantite_decimal,
                    prix_unitaire_decimal,
                    isin=isin
                )

                if existing_transaction:
                    nombre_doublons += 1
                    erreurs.append({
                        'ligne': row_idx,
                        'type': 'doublon',
                        'erreur': 'Transaction déjà importée (doublon détecté)'
                    })
                else:
                    file_storage.create_transaction(
                        portfolio['id'],
                        signaletique_id,
                        date_transaction,
                        type_operation,
                        quantite_decimal,
                        prix_unitaire_decimal,
                        devise,
                        isin=isin
                    )
                    
                    nombre_succes += 1
                
            except Exception as e:
                nombre_erreurs += 1
                erreurs.append({
                    'ligne': row_idx,
                    'type': 'erreur',
                    'erreur': str(e)
                })
        
        # Générer le fichier CSV pour les ISINs inconnus
        csv_file_url = None
        if isins_inconnus:
            # Créer le dossier media/imports s'il n'existe pas
            import_dir = os.path.join(settings.MEDIA_ROOT, 'imports')
            os.makedirs(import_dir, exist_ok=True)
            
            # Nom du fichier avec timestamp
            from datetime import datetime as dt
            timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
            csv_filename = f'signaletique_isins_inconnus_{timestamp}.csv'
            csv_filepath = os.path.join(import_dir, csv_filename)
            
            # Créer le fichier CSV avec les en-têtes du template
            with open(csv_filepath, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # En-têtes basés sur le template
                headers_csv = [
                    "Type d'instr", "Isin", "Classe d'actifs", "Nom", "Symbole", 
                    "Banques Dispo", "Devise", "Taux", "Date de fin", "Qualité credit",
                    "TER", "Cap/Dis", "ESG", "Replication", "Taille du Fonds", 
                    "Positions", "Couverture de change", "USA", "Japon", "Grande Bretagne",
                    "Canada", "Pays Emergeants Hors Chine et Japon", "Australie", "Suède",
                    "Suisse", "Chine", "Israel", "Allemagne", "Nouvelle Zelande", 
                    "Pays-Bas", "Irlande", "Espagne", "Italie", "France", "Autre Pays",
                    "Etats", "Industrie", "Finance", "Consommation Cyclique", "Technologie",
                    "Santé", "Consommation Defensive", "Communication", "Immobilier",
                    "Matières Premières", "Energie", "Service Publiques", 
                    "Services de consommation", "Autre Secteur", "Etats2", 
                    "Banque Emetteur", "Entreprises", "Autre Emetteur", "CodeBank"
                ]
                writer.writerow(headers_csv)
                
                # Ajouter une ligne pour chaque ISIN inconnu
                for isin in sorted(isins_inconnus):
                    row = [''] * len(headers_csv)
                    row[1] = isin  # Colonne "Isin"
                    row[3] = f'À compléter pour {isin}'  # Colonne "Nom"
                    writer.writerow(row)
            
            # URL relative pour le téléchargement
            csv_file_url = f'/media/imports/{csv_filename}'
        
        # Mettre à jour le log
        import_log.nombre_lignes = ws.max_row - 1
        import_log.nombre_succes = nombre_succes
        import_log.nombre_erreurs = nombre_erreurs
        import_log.erreurs = erreurs if erreurs else None
        import_log.statut = 'termine' if nombre_erreurs == 0 else 'termine_avec_erreurs'
        import_log.save()
        
        response_data = {
            'success': True,
            'message': f'Import terminé avec succès',
            'details': {
                'fichier': file.name,
                'colonnes_detectees': headers,
                'lignes_totales': ws.max_row - 1,
                'succes': nombre_succes,
                'doublons': nombre_doublons,
                'erreurs': nombre_erreurs,
                'liste_erreurs': erreurs[:10] if erreurs else []
            }
        }
        
        # Ajouter l'info sur les ISINs inconnus
        if isins_inconnus:
            response_data['isins_inconnus'] = {
                'count': len(isins_inconnus),
                'liste': sorted(list(isins_inconnus)),
                'csv_file_url': csv_file_url
            }
        
        return Response(response_data)
        
    except Exception as e:
        import_log.statut = 'erreur'
        import_log.erreurs = [{'erreur_generale': str(e)}]
        import_log.save()
        
        return Response(
            {'error': f'Erreur lors de l\'import: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET', 'POST'])
def list_transactions(request):
    """Liste ou crée des transactions"""

    if request.method == 'POST':
        data = request.data

        portfolio_id = data.get('portfolio_id')
        isin = data.get('isin', '').strip().upper()
        date_transaction = data.get('date')
        type_operation = data.get('type_operation', '').strip().upper()
        quantite = data.get('quantite')
        prix_unitaire = data.get('prix_unitaire')
        devise = data.get('devise', 'EUR')

        if not all([portfolio_id, isin, date_transaction, type_operation, quantite, prix_unitaire]):
            return Response(
                {'error': 'Champs obligatoires manquants: portfolio_id, isin, date, type_operation, quantite, prix_unitaire'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if type_operation not in ['ACHAT', 'VENTE']:
            return Response({'error': "type_operation doit être ACHAT ou VENTE"}, status=status.HTTP_400_BAD_REQUEST)

        portfolio = file_storage.get_portfolio_by_id(int(portfolio_id))
        if not portfolio:
            return Response({'error': 'Portefeuille non trouvé'}, status=status.HTTP_404_NOT_FOUND)

        try:
            sig_dict = file_storage.get_signaletique_by_isin(isin)
            if not sig_dict:
                return Response({'error': f"ISIN {isin} introuvable dans la signalétique"}, status=status.HTTP_404_NOT_FOUND)
            signaletique_id = sig_dict['id']
        except Exception:
            return Response({'error': f"ISIN {isin} introuvable dans la signalétique"}, status=status.HTTP_404_NOT_FOUND)

        try:
            quantite_decimal = Decimal(str(quantite))
            prix_decimal = Decimal(str(prix_unitaire))
        except Exception:
            return Response({'error': 'quantite et prix_unitaire doivent être des nombres valides'}, status=status.HTTP_400_BAD_REQUEST)

        transaction = file_storage.create_transaction(
            int(portfolio_id),
            signaletique_id,
            date_transaction,
            type_operation,
            quantite_decimal,
            prix_decimal,
            devise,
            isin=isin
        )

        return Response(transaction, status=status.HTTP_201_CREATED)

    # GET
    portfolio_id = request.GET.get('portfolio_id')

    if portfolio_id:
        transactions = file_storage.get_transactions_by_portfolio(int(portfolio_id))
    else:
        transactions = file_storage.get_all_transactions()

    # Enrichir les transactions avec les infos de signalétique
    enriched_transactions = []
    for transaction in transactions:
        sig = file_storage.get_signaletique_for_transaction(transaction)
        portfolio = file_storage.get_portfolio_by_id(transaction['portfolio_id'])

        enriched_transaction = transaction.copy()
        if sig:
            enriched_transaction['signaletique'] = {
                'id': sig['id'],
                'code': sig.get('code', ''),
                'isin': sig.get('isin', ''),
                'titre': sig.get('titre', '')
            }
        if portfolio:
            enriched_transaction['portfolio'] = {
                'id': portfolio['id'],
                'name': portfolio['name']
            }

        enriched_transactions.append(enriched_transaction)

    return Response(enriched_transactions)


@api_view(['DELETE'])
def delete_transaction(request, pk):
    """Supprime une transaction"""
    if file_storage.delete_transaction(int(pk)):
        return Response(
            {'success': True, 'message': 'Transaction supprimée'},
            status=status.HTTP_204_NO_CONTENT
        )
    
    return Response(
        {'error': 'Transaction non trouvée'},
        status=status.HTTP_404_NOT_FOUND
    )


def is_obligation(categorie, type_instrument=''):
    """
    Détermine si un actif est une obligation.
    Pour les obligations, le prix est en pourcentage (100 = 1).
    """
    categorie_lower = (categorie or '').lower()
    type_lower = (type_instrument or '').lower()
    return 'obligation' in categorie_lower or 'obligation' in type_lower


def get_prix_reel(prix_unitaire, categorie, type_instrument=''):
    """
    Retourne le prix réel en tenant compte du fait que les obligations sont en pourcentage.
    """
    if is_obligation(categorie, type_instrument):
        return prix_unitaire / Decimal('100')
    return prix_unitaire


@api_view(['GET'])
def portfolio_fifo_analysis(request, pk):
    """Analyse FIFO d'un portefeuille : positions actuelles et P&L réalisé"""
    portfolio = file_storage.get_portfolio_by_id(int(pk))
    
    if not portfolio:
        return Response(
            {'error': 'Portefeuille non trouvé'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Récupérer toutes les transactions triées par date
    all_transactions = file_storage.get_transactions_by_portfolio(int(pk))
    
    # Trier les transactions par date et id
    transactions_sorted = sorted(all_transactions, key=lambda x: (x['date'], x['id']))
    
    # Structure pour stocker les lots d'achat (FIFO)
    # positions[isin] = [(date, quantite, prix_unitaire), ...]
    positions = {}
    
    # P&L réalisé par titre
    realized_pnl = {}
    
    # Historique des transactions traitées
    transaction_details = []
    
    for transaction in transactions_sorted:
        # Récupérer la signalétique
        sig = file_storage.get_signaletique_for_transaction(transaction)
        if not sig:
            continue

        isin = sig.get('isin') or sig.get('code')
        titre = sig.get('titre', '')
        categorie = sig.get('categorie_text') or 'Non classé'

        # Extraire le type d'instrument depuis donnees_supplementaires
        type_instrument = ''
        ds = sig.get('donnees_supplementaires') or {}
        if "Type d'instr" in ds:
            type_instrument = ds["Type d'instr"] or ''
        
        # Convertir les valeurs string en Decimal
        quantite = Decimal(str(transaction['quantite']))
        prix_unitaire = Decimal(str(transaction['prix_unitaire']))
        
        if transaction['type_operation'] == 'ACHAT':
            # Ajouter un lot d'achat
            if isin not in positions:
                positions[isin] = {
                    'titre': titre,
                    'categorie': categorie,
                    'type_instrument': type_instrument,
                    'signaletique_id': sig['id'],
                    'donnees_supplementaires': ds,
                    'lots': [],
                    'quantite_totale': Decimal('0')
                }
            
            positions[isin]['lots'].append({
                'date': transaction['date'],
                'quantite': quantite,
                'prix_unitaire': prix_unitaire,
                'devise': transaction['devise']
            })
            positions[isin]['quantite_totale'] += quantite
            
            # Calculer le montant (pour obligations: prix en %, donc diviser par 100)
            prix_reel = get_prix_reel(prix_unitaire, categorie, type_instrument)
            montant = quantite * prix_reel
            
            transaction_details.append({
                'id': transaction['id'],
                'date': transaction['date'],
                'type': 'ACHAT',
                'titre': titre,
                'categorie': categorie,
                'type_instrument': type_instrument,
                'quantite': float(quantite),
                'prix_unitaire': float(prix_unitaire),
                'montant': float(montant),
                'devise': transaction['devise']
            })
            
        elif transaction['type_operation'] == 'VENTE':
            # Consommer les lots d'achat en FIFO
            if isin not in positions or positions[isin]['quantite_totale'] < quantite:
                # Vente à découvert ou erreur
                prix_reel = get_prix_reel(prix_unitaire, categorie, type_instrument)
                montant = quantite * prix_reel
                
                transaction_details.append({
                    'id': transaction['id'],
                    'date': transaction['date'],
                    'type': 'VENTE',
                    'titre': titre,
                    'categorie': categorie,
                    'quantite': float(quantite),
                    'prix_unitaire': float(prix_unitaire),
                    'montant': float(montant),
                    'devise': transaction['devise'],
                    'erreur': 'Vente sans achat correspondant'
                })
                continue
            
            quantite_a_vendre = quantite
            pnl_transaction = Decimal('0')
            lots_consommes = []
            
            while quantite_a_vendre > 0 and positions[isin]['lots']:
                lot = positions[isin]['lots'][0]
                
                if lot['quantite'] <= quantite_a_vendre:
                    # Consommer tout le lot
                    quantite_vendue = lot['quantite']
                    # Pour obligations: prix en %, donc calculer avec prix réels
                    prix_vente_reel = get_prix_reel(prix_unitaire, categorie, type_instrument)
                    prix_achat_reel = get_prix_reel(lot['prix_unitaire'], categorie, type_instrument)
                    pnl_lot = quantite_vendue * (prix_vente_reel - prix_achat_reel)
                    pnl_transaction += pnl_lot
                    
                    lots_consommes.append({
                        'quantite': float(quantite_vendue),
                        'prix_achat': float(lot['prix_unitaire']),
                        'prix_vente': float(prix_unitaire),
                        'pnl': float(pnl_lot)
                    })
                    
                    quantite_a_vendre -= quantite_vendue
                    positions[isin]['quantite_totale'] -= quantite_vendue
                    positions[isin]['lots'].pop(0)
                    
                else:
                    # Consommer partiellement le lot
                    quantite_vendue = quantite_a_vendre
                    # Pour obligations: prix en %, donc calculer avec prix réels
                    prix_vente_reel = get_prix_reel(prix_unitaire, categorie, type_instrument)
                    prix_achat_reel = get_prix_reel(lot['prix_unitaire'], categorie, type_instrument)
                    pnl_lot = quantite_vendue * (prix_vente_reel - prix_achat_reel)
                    pnl_transaction += pnl_lot
                    
                    lots_consommes.append({
                        'quantite': float(quantite_vendue),
                        'prix_achat': float(lot['prix_unitaire']),
                        'prix_vente': float(prix_unitaire),
                        'pnl': float(pnl_lot)
                    })
                    
                    lot['quantite'] -= quantite_vendue
                    positions[isin]['quantite_totale'] -= quantite_vendue
                    quantite_a_vendre = Decimal('0')
            
            # Enregistrer le P&L réalisé
            if isin not in realized_pnl:
                realized_pnl[isin] = {
                    'titre': titre,
                    'categorie': categorie,
                    'type_instrument': type_instrument,
                    'pnl_total': Decimal('0'),
                    'ventes': []
                }
            
            realized_pnl[isin]['pnl_total'] += pnl_transaction
            realized_pnl[isin]['ventes'].append({
                'date': transaction['date'],
                'quantite': float(quantite),
                'prix_vente': float(prix_unitaire),
                'pnl': float(pnl_transaction),
                'lots_consommes': lots_consommes
            })
            
            # Calculer le montant (pour obligations: prix en %, donc diviser par 100)
            prix_reel = get_prix_reel(prix_unitaire, categorie, type_instrument)
            montant = quantite * prix_reel
            
            transaction_details.append({
                'id': transaction['id'],
                'date': transaction['date'],
                'type': 'VENTE',
                'titre': titre,
                'categorie': categorie,
                'type_instrument': type_instrument,
                'quantite': float(quantite),
                'prix_unitaire': float(prix_unitaire),
                'montant': float(montant),
                'devise': transaction['devise'],
                'pnl': float(pnl_transaction),
                'lots_consommes': lots_consommes
            })
    
    # Nettoyer les positions vides
    positions_actuelles = []
    for isin, data in positions.items():
        if data['quantite_totale'] > 0:
            # Calculer le prix moyen pondéré des lots restants
            # Pour obligations: utiliser prix réel (prix/100)
            categorie = data.get('categorie', 'Non classé')
            type_instrument = data.get('type_instrument', '')
            
            valeur_totale = Decimal('0')
            for lot in data['lots']:
                prix_reel = get_prix_reel(lot['prix_unitaire'], categorie, type_instrument)
                valeur_totale += lot['quantite'] * prix_reel
            
            prix_moyen = valeur_totale / data['quantite_totale'] if data['quantite_totale'] > 0 else Decimal('0')
            
            positions_actuelles.append({
                'isin': isin,
                'titre': data['titre'],
                'categorie': data.get('categorie', 'Non classé'),
                'type_instrument': data.get('type_instrument', ''),
                'signaletique_id': data.get('signaletique_id'),
                'donnees_supplementaires': data.get('donnees_supplementaires', {}),
                'quantite': float(data['quantite_totale']),
                'prix_moyen': float(prix_moyen),
                'valeur': float(valeur_totale),
                'devise': data['lots'][0]['devise'] if data['lots'] else 'EUR',
                'lots': [
                    {
                        'date': lot['date'],
                        'quantite': float(lot['quantite']),
                        'prix_unitaire': float(lot['prix_unitaire'])
                    }
                    for lot in data['lots']
                ]
            })
    
    # Calculer le P&L total réalisé
    pnl_total_realise = sum(data['pnl_total'] for data in realized_pnl.values())
    
    # Formater le P&L réalisé
    pnl_realise_details = [
        {
            'isin': isin,
            'titre': data['titre'],
            'categorie': data.get('categorie', 'Non classé'),
            'type_instrument': data.get('type_instrument', ''),
            'pnl_total': float(data['pnl_total']),
            'ventes': data['ventes']
        }
        for isin, data in realized_pnl.items()
    ]
    
    return Response({
        'portfolio': {
            'id': portfolio['id'],
            'name': portfolio['name'],
            'description': portfolio.get('description', ''),
            'type_compte': portfolio.get('type_compte', ''),
            'courtier': portfolio.get('courtier', ''),
            'devise': portfolio.get('devise', 'EUR'),
            'date_ouverture': portfolio.get('date_ouverture', ''),
            'couleur': portfolio.get('couleur', ''),
            'date_creation': portfolio.get('date_creation', ''),
        },
        'positions_actuelles': positions_actuelles,
        'pnl_realise': {
            'total': float(pnl_total_realise),
            'par_titre': pnl_realise_details
        },
        'transactions': transaction_details,
        'statistiques': {
            'nombre_transactions': len(transaction_details),
            'nombre_titres_en_portefeuille': len(positions_actuelles),
            'nombre_titres_vendus': len(pnl_realise_details)
        }
    })
