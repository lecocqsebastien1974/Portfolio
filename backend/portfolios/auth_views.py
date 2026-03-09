from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status


@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """Connexion utilisateur"""
    username = request.data.get('username')
    password = request.data.get('password')
    
    if not username or not password:
        return Response(
            {'error': 'Nom d\'utilisateur et mot de passe requis'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = authenticate(request, username=username, password=password)
    
    if user is not None:
        login(request, user)
        return Response({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser
            }
        })
    else:
        return Response(
            {'error': 'Identifiants invalides'},
            status=status.HTTP_401_UNAUTHORIZED
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    """Déconnexion utilisateur"""
    logout(request)
    return Response({'success': True, 'message': 'Déconnexion réussie'})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    """Récupérer l'utilisateur connecté"""
    user = request.user
    return Response({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'is_staff': user.is_staff,
        'is_superuser': user.is_superuser
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser])
def import_users(request):
    """Restaure les utilisateurs depuis un fichier Excel de sauvegarde.
    Colonnes attendues : username, email, first_name, last_name,
                         is_staff, is_superuser, is_active, password_hash
    Seul un superuser peut utiliser cet endpoint.
    """
    if not request.user.is_superuser:
        return Response({'error': 'Réservé aux superusers'}, status=status.HTTP_403_FORBIDDEN)

    if 'file' not in request.FILES:
        return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    if not file.name.endswith(('.xlsx', '.xls')):
        return Response({'error': 'Format non supporté. Utilisez .xlsx ou .xls'}, status=status.HTTP_400_BAD_REQUEST)

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]

        crees = 0
        mis_a_jour = 0
        erreurs = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {h: v for h, v in zip(headers, row)}
            username = str(row_data.get('username') or '').strip()
            if not username:
                continue

            password_hash = str(row_data.get('password_hash') or '').strip()
            if not password_hash:
                erreurs.append({'ligne': row_idx, 'erreur': 'password_hash manquant pour %s' % username})
                continue

            try:
                user, created = User.objects.get_or_create(
                    username=username,
                    defaults={
                        'email': str(row_data.get('email') or ''),
                        'first_name': str(row_data.get('first_name') or ''),
                        'last_name': str(row_data.get('last_name') or ''),
                        'is_staff': bool(row_data.get('is_staff')),
                        'is_superuser': bool(row_data.get('is_superuser')),
                        'is_active': row_data.get('is_active') != False,
                    }
                )
                # Restaurer le hash directement (sans re-hachage)
                user.password = password_hash
                if not created:
                    user.email = str(row_data.get('email') or '')
                    user.first_name = str(row_data.get('first_name') or '')
                    user.last_name = str(row_data.get('last_name') or '')
                    user.is_staff = bool(row_data.get('is_staff'))
                    user.is_superuser = bool(row_data.get('is_superuser'))
                    user.is_active = row_data.get('is_active') != False
                user.save()
                if created:
                    crees += 1
                else:
                    mis_a_jour += 1
            except Exception as e:
                erreurs.append({'ligne': row_idx, 'erreur': str(e)})

        return Response({
            'success': True,
            'details': {
                'crees': crees,
                'mis_a_jour': mis_a_jour,
                'erreurs': len(erreurs),
                'liste_erreurs': erreurs[:10]
            }
        })
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([AllowAny])
def check_auth(request):
    """Vérifier si l'utilisateur est authentifié"""
    if request.user.is_authenticated:
        return Response({
            'authenticated': True,
            'user': {
                'id': request.user.id,
                'username': request.user.username,
                'email': request.user.email,
                'first_name': request.user.first_name,
                'last_name': request.user.last_name,
                'is_staff': request.user.is_staff,
                'is_superuser': request.user.is_superuser
            }
        })
    else:
        return Response({'authenticated': False})
