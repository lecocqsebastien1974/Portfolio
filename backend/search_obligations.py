#!/usr/bin/env python
import openpyxl

target_isins = ['US298785HM16', 'NZGOVDT528C6', 'NZGOVDT526C0', 'US045167DR18', 'US91282CAD39', 'IE00BYZTVV78']

files_to_check = [
    '/app/signaletique14022026.xlsx',
    '/app/Signalétique020326.xlsx',
    '/app/Signaletique_main.xlsx'
]

print("Recherche des ISINs obligations:")
print("="*80)

for filepath in files_to_check:
    try:
        wb = openpyxl.load_workbook(filepath)
        print(f"\n📁 Fichier: {filepath}")
        print(f"Feuilles: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            if ws.max_row < 2:
                continue
            
            headers = [cell.value for cell in ws[1]]
            
            # Trouver colonne ISIN
            isin_col_idx = None
            for idx, h in enumerate(headers):
                if h and 'isin' in str(h).lower():
                    isin_col_idx = idx
                    break
            
            if isin_col_idx is None:
                continue
            
            found_in_sheet = False
            
            # Chercher les ISINs
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                isin = row[isin_col_idx] if isin_col_idx < len(row) else None
                if isin and str(isin).strip() in target_isins:
                    if not found_in_sheet:
                        print(f"\n  📄 Feuille: {sheet_name}")
                        found_in_sheet = True
                    
                    print(f"\n    ✅ ISIN: {isin} (ligne {row_idx})")
                    print("    " + "-" * 70)
                    for idx, (header, value) in enumerate(zip(headers, row)):
                        if value and str(value).strip():
                            print(f"      {header}: {value}")
                    print("    " + "-" * 70)
    
    except Exception as e:
        print(f"  ❌ Erreur: {e}")

print("\n\n" + "="*80)
print("Recherche terminée")
print("="*80)
