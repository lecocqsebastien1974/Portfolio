#!/usr/bin/env python
import openpyxl

# Charger le fichier Excel
wb = openpyxl.load_workbook('/app/achat_ETF_et_fonds_oblig2.xlsx')
ws = wb.active

# Afficher les en-têtes
headers = [cell.value for cell in ws[1]]
print("En-têtes:")
for idx, h in enumerate(headers, 1):
    print(f"  {idx}. {h}")

print("\n" + "="*80)
print("Premières lignes:")
print("="*80)

# Afficher les premières lignes
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=10, values_only=True), 2):
    print(f"\nLigne {row_idx}:")
    for idx, (header, value) in enumerate(zip(headers, row), 1):
        if value:
            print(f"  {header}: {value}")

# Chercher les ISINs spécifiques
print("\n" + "="*80)
print("Recherche des ISINs à compléter:")
print("="*80)

target_isins = ['FR0010149179', 'GB00B00FHZ82', 'LU0171289902']
isin_col_idx = None

# Trouver l'index de la colonne ISIN
for idx, header in enumerate(headers):
    if header and 'isin' in str(header).lower():
        isin_col_idx = idx
        print(f"Colonne ISIN trouvée: '{headers[idx]}' (index {idx})")
        break

if isin_col_idx is not None:
    for row in ws.iter_rows(min_row=2, values_only=True):
        isin = row[isin_col_idx] if isin_col_idx < len(row) else None
        if isin and str(isin).strip() in target_isins:
            print(f"\n✓ ISIN trouvé: {isin}")
            for idx, (header, value) in enumerate(zip(headers, row)):
                if value:
                    print(f"  {header}: {value}")
