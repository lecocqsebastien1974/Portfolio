#!/usr/bin/env python
import openpyxl

target_isins = ['FR0010149179', 'GB00B00FHZ82', 'LU0171289902']

# Charger le fichier
wb = openpyxl.load_workbook('/app/Signaletique_main.xlsx')

print(f"Feuilles: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    
    # Trouver colonne ISIN
    isin_col_idx = None
    for idx, h in enumerate(headers):
        if h and ('isin' in str(h).lower() or 'code' in str(h).lower()):
            isin_col_idx = idx
            break
    
    if isin_col_idx is None:
        continue
    
    print(f"\n{'='*80}")
    print(f"Feuille: {sheet_name}")
    print(f"Colonne ISIN/Code: '{headers[isin_col_idx]}' (idx {isin_col_idx})")
    print('='*80)
    
    # Chercher les ISINs
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        isin = row[isin_col_idx] if isin_col_idx < len(row) else None
        if isin:
            isin_str = str(isin).strip()
            if isin_str in target_isins:
                print(f"\n✓✓✓ TROUVÉ ligne {row_idx}: {isin_str}")
                print("-" * 60)
                for idx, (header, value) in enumerate(zip(headers, row)):
                    if value and str(value).strip():
                        print(f"  {header}: {value}")
                print("-" * 60)
