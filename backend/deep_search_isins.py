#!/usr/bin/env python
import openpyxl

target_isins = ['FR0010149179', 'GB00B00FHZ82', 'LU0171289902']

files_to_check = [
    '/app/Signalétique020326.xlsx',
    '/app/Signaletique_main.xlsx'
]

for filepath in files_to_check:
    try:
        wb = openpyxl.load_workbook(filepath)
        print(f"\n{'#'*80}")
        print(f"Fichier: {filepath}")
        print(f"Feuilles: {wb.sheetnames}")
        print('#'*80)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            if ws.max_row < 2:
                continue
                
            headers = [cell.value for cell in ws[1]]
            
            # Chercher toutes les colonnes qui pourraient contenir un ISIN
            potential_isin_cols = []
            for idx, h in enumerate(headers):
                if h and any(keyword in str(h).lower() for keyword in ['isin', 'code', 'titre', 'valeur', 'security']):
                    potential_isin_cols.append((idx, h))
            
            if not potential_isin_cols:
                continue
            
            print(f"\n  Feuille: {sheet_name}")
            print(f"  Colonnes potentielles: {[h for _,h in potential_isin_cols]}")
            
            # Chercher dans toutes les lignes
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                row_has_target = False
                target_found = None
                
                for col_idx, col_name in potential_isin_cols:
                    cell_value = row[col_idx] if col_idx < len(row) else None
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if cell_str in target_isins:
                            row_has_target = True
                            target_found = cell_str
                            break
                
                if row_has_target:
                    print(f"\n    ✓✓✓ TROUVÉ: {target_found} (ligne {row_idx})")
                    print("    " + "-" * 70)
                    for idx, (header, value) in enumerate(zip(headers, row)):
                        if value and str(value).strip():
                            print(f"      {header}: {value}")
                    print("    " + "-" * 70)
    
    except Exception as e:
        print(f"  ❌ Erreur: {e}")

print("\n\n" + "="*80)
print("Recherche terminée")
print("="*80)
