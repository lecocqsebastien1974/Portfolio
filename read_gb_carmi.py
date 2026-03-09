#!/usr/bin/env python
import openpyxl

# Charger le fichier
wb = openpyxl.load_workbook('/app/Signalétique Template gb et carmi.xlsx')

print(f"Feuilles disponibles: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    print(f"{'='*80}")
    print(f"Feuille: {sheet_name}")
    print('='*80)
    
    if ws.max_row < 2:
        print("  Feuille vide\n")
        continue
    
    # Obtenir les en-têtes
    headers = [cell.value for cell in ws[1]]
    print(f"\nEn-têtes: {headers[:10]}\n")
    
    # Trouver colonne ISIN
    isin_col_idx = None
    for idx, h in enumerate(headers):
        if h and 'isin' in str(h).lower():
            isin_col_idx = idx
            print(f"Colonne ISIN trouvée: '{h}' (index {idx})\n")
            break
    
    if isin_col_idx is None:
        print("⚠️  Pas de colonne ISIN trouvée\n")
        continue
    
    # Afficher toutes les lignes avec leurs données
    print("Données:")
    print("-" * 80)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        isin = row[isin_col_idx] if isin_col_idx < len(row) else None
        if isin and str(isin).strip():
            print(f"\nLigne {row_idx} - ISIN: {isin}")
            for idx, (header, value) in enumerate(zip(headers, row)):
                if value and str(value).strip():
                    if isinstance(value, str) and value.startswith('='):
                        continue
                    print(f"  {header}: {value}")
    
    print("\n")
